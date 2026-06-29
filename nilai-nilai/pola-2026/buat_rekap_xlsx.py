#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Membuat FILE BARU rekap nilai berbasis template resmi
POLA_PAIK6808_2025_2_A.xlsx, TANPA mengubah file aslinya.

Output: POLA_PAIK6808_2025_2_A-rekap-nilai.xlsx

Isi: struktur sheet resmi dipertahankan, lalu tiap komponen nilai
ditampilkan TERPISAH per kolom di samping (kolom N = pemisah):
    O: Classification Task (Intermediate)   [0-10]
    P: Evaluation Metric Quiz               [0-10]
    Q: Manual Classification Task (Beginner)[0-10]
    R: Oral                                 [0-10]
    S: Arsitektur                           [0-10]
    T: Proyek (Presentasi + Keaktifan)      [0-100]
Nilai apa adanya (raw), tidak dirata-rata. Kolom resmi F & H dibiarkan kosong.
"""
import os, openpyxl
from openpyxl.styles import Font, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "POLA_PAIK6808_2025_2_A.xlsx")            # template asli (tidak diubah)
OUT = os.path.join(BASE, "POLA_PAIK6808_2025_2_A-rekap-nilai.xlsx")  # FILE BARU

# NIM -> (ClassificationTask, EvaluationMetricQuiz, ManualClassification, Oral, Arsitektur, ProyekPK)
DATA = {
 '24060119120041': ('', '2.50', '', '6.50', '', '0'),
 '24060123120030': ('9.67', '5.00', '10.00', '9.30', '9.00', '95'),
 '24060123120029': ('9.00', '10.00', '3.33', '8.70', '8.80', '80'),
 '24060123130093': ('10.00', '7.50', '10.00', '6.00', '9.40', '92'),
 '24060123130084': ('8.00', '0.00', '3.33', '9.20', '8.70', '80'),
 '24060123140201': ('8.33', '10.00', '3.33', '8.50', '8.70', '80'),
 '24060123130058': ('10.00', '10.00', '3.33', '7.50', '8.70', '80'),
 '24060123120002': ('10.00', '10.00', '8.33', '8.00', '9.30', '80'),
 '24060123130051': ('', '', '', '', '', '82'),
 '24060123130079': ('9.33', '10.00', '6.67', '9.20', '8.90', '-10'),
 '24060123140175': ('9.00', '10.00', '5.00', '8.50', '9.00', '80'),
 '24060123130073': ('', '', '', '', '', '85'),
 '24060123140045': ('9.00', '', '8.33', '8.50', '8.90', '0'),
 '24060123120040': ('9.67', '7.50', '10.00', '9.00', '9.00', '87'),
 '24060123130080': ('9.00', '7.50', '8.33', '7.80', '8.30', '82'),
 '24060123120016': ('', '', '', '', '', '82'),
 '24060123140147': ('7.67', '10.00', '5.00', '', '', '80'),
 '24060123140111': ('10.00', '10.00', '8.33', '8.30', '9.50', '80'),
 '24060123130098': ('6.00', '7.50', '8.33', '8.20', '8.50', '80'),
 '24060123140211': ('', '0.00', '', '7.30', '9.40', '-10'),
 '24060123130094': ('8.67', '0.00', '6.67', '8.50', '8.60', '87'),
 '24060123120028': ('7.67', '0.00', '6.67', '8.30', '9.20', '87'),
 '24060123130112': ('10.00', '10.00', '10.00', '9.00', '9.50', '0'),
 '24060123130110': ('9.67', '10.00', '8.33', '8.00', '8.90', '0'),
 '24060123140197': ('9.33', '2.50', '8.33', '7.80', '9.00', '90'),
 '24060123140151': ('8.33', '10.00', '8.33', '7.50', '8.80', '90'),
 '24060120130039': ('', '', '', '', '', '-10'),
 '24060123120009': ('8.67', '5.00', '3.33', '8.30', '8.70', '87'),
 '24060123140150': ('9.67', '0.00', '10.00', '6.50', '8.90', '80'),
 '24060123130101': ('', '', '', '', '', '-10'),
 '24060123130117': ('9.67', '10.00', '8.33', '9.00', '9.00', '90'),
 '24060123140179': ('9.67', '10.00', '10.00', '8.50', '9.30', '0'),
 '24060123140152': ('', '', '0.00', '', '8.60', '0'),
 '24060123130081': ('9.33', '10.00', '10.00', '7.50', '9.20', '0'),
 '24060123140166': ('9.00', '0.00', '5.00', '7.00', '9.00', '90'),
 '24060123120032': ('6.67', '7.50', '6.67', '9.40', '8.80', '0'),
 '24060123140204': ('9.33', '10.00', '8.33', '9.30', '8.90', '90'),
 '24060123140139': ('', '', '', '', '', '87'),
 '24060123130086': ('8.00', '', '6.67', '', '8.60', '80'),
 '24060123130107': ('6.33', '5.00', '8.33', '7.50', '', '-10'),
 '24060123130106': ('8.00', '7.50', '6.67', '', '9.50', '0'),
 '24060123120023': ('9.33', '5.00', '8.33', '8.00', '8.70', '0'),
 '24060123120038': ('', '', '', '', '8.80', '0'),
 '24060123140148': ('9.67', '7.50', '10.00', '8.00', '8.90', '80'),
 '24060123130067': ('8.00', '', '8.33', '8.50', '8.80', '0'),
 '24060123140142': ('7.00', '', '3.33', '7.00', '9.00', '0'),
 '24060123130100': ('', '7.50', '', '8.00', '8.60', '77'),
 '24060122140184': ('', '0.00', '', '4.00', '8.50', '0'),
 '24060123120010': ('8.00', '10.00', '5.00', '8.20', '9.00', '82'),
 '24060123130114': ('9.33', '5.00', '10.00', '7.30', '7.80', '0'),
}

def num(s):
    s = (s or "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None

SIDE = [
    (15, "Classification Task (Intermediate)"),
    (16, "Evaluation Metric Quiz"),
    (17, "Manual Classification Task (Beginner)"),
    (18, "Oral"),
    (19, "Arsitektur"),
    (20, "Proyek (Presentasi + Keaktifan)"),
]
HEADER_ROW = 7

wb = openpyxl.load_workbook(SRC)   # buka template asli (read; tidak disimpan ke SRC)
ws = wb["Worksheet"]

ws.cell(HEADER_ROW, 14).value = "|"
for col, title in SIDE:
    c = ws.cell(HEADER_ROW, col)
    c.value = title
    c.font = Font(bold=True)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[c.column_letter].width = 16

filled = 0
for r in range(8, ws.max_row + 1):
    nim = ws.cell(r, 1).value
    if nim is None:
        continue
    nim = str(nim).strip()
    if nim not in DATA:
        continue
    ct, quiz, manual, oral, arsi, ppk = DATA[nim]
    for (col, _), v in zip(SIDE, [num(ct), num(quiz), num(manual), num(oral), num(arsi), num(ppk)]):
        ws.cell(r, col).value = v
    ws.cell(r, 14).value = "|"
    filled += 1

wb.save(OUT)   # simpan ke FILE BARU
print(f"File baru dibuat: {os.path.basename(OUT)}")
print(f"File asli TIDAK diubah: {os.path.basename(SRC)}")
print(f"Baris terisi: {filled}")
