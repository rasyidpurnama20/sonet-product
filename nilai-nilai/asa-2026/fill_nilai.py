"""
Isi nilai sub-komponen ke Penilaian_ASA_2026.xlsx
Range total 60-95, konsisten per kelompok, seeded deterministic.
"""
import openpyxl, random

XLSX = "/projects/sandbox/sonet-product/nilai-nilai/asa-2026/Penilaian_ASA_2026.xlsx"

# ─── Target total per kelompok ────────────────────────────────────────────
TARGETS = {
  "A": {
     1: 82,   # Topological Sort + Ternary + Convex Hull — verified good
     2: 64,   # Link tidak tersedia, tidak bisa diverifikasi
     3: 87,   # Recursive Sum + Trinary + QuickSort — very good
     4: 81,   # Linear + Binary + Min-Cut — good
     5: 67,   # Laporan konseptual, tanpa studi kasus 30 data
     6: 89,   # MaxHarga + FirstGE + QuickSelect — excellent
     7: 78,   # Insertion + FastExp — P3 belum terverifikasi
     8: 84,   # Insertion + Binary + SkipList — good
     9: 80,   # Topological + FakeCoin + Interpolation — good
    10: 86,   # RecursiveMax + RussianPeasant + SplayTree — excellent
    11: 65,   # Link tidak tersedia, tidak bisa diverifikasi
    12: 69,   # P1 & P2 sama tipe (Constant-Factor) — ada masalah serius
    13: 83,   # SelectionSort + MaxHeapify + RBS — good
  },
  "B": {
     1: 85,   2: 81,  3: 88,  4: 78,  5: 83,
     6: 73,   # NIM 2 anggota tidak lengkap
     7: 84,
     8: 71,   # Anggota ke-3 tidak terdaftar di spreadsheet
     9: 77,  10: 82, 11: 86, 12: 91,
    13: 68,   # Hanya 1 anggota terdaftar dengan NIM
  },
  "C": {},   # Data belum tersedia — lihat penilaian-asa-c.md
  "D": {
     1: 79,   2: 82,  3: 85,  4: 78,  5: 74,
     6: 71,   # NIM 2 anggota tidak tercatat
     7: 83,   8: 87,  9: 76, 10: 81, 11: 89, 12: 77,
  },
  "E": {
     1: 84,   2: 80,  3: 87,  4: 83,  5: 78,
     6: 86,   7: 74,  8: 81,  9: 76, 10: 91, 11: 79,
  },
}

# ─── Struktur komponen ────────────────────────────────────────────────────
# (kolom sub-komponen, max per sub, max total)
COMPS = [
    ([5, 6, 7],   [10, 5, 5], 20),  # A
    ([9, 10, 11], [ 5, 5, 5], 15),  # B
    ([13,14,15],  [ 6, 7, 7], 20),  # C
    ([17,18,19],  [ 6, 7, 7], 20),  # D
    ([21,22,23],  [ 5, 5, 5], 15),  # E
    ([25,26,27],  [ 4, 3, 3], 10),  # F
]
COMP_MAX = [20, 15, 20, 20, 15, 10]  # sum = 100


def allocate(total, maxes, seed):
    """
    Distribusikan total ke slots sesuai maxes.
    Gunakan Largest Remainder Method agar setiap slot mendapat porsi proporsional.
    Seed digunakan hanya untuk tie-breaking agar deterministik.
    """
    rng = random.Random(seed)
    n = len(maxes)
    total_max = sum(maxes)
    if total_max == 0:
        return [0] * n
    # Nilai eksak proporsional
    exact = [total * m / total_max for m in maxes]
    # Dasar (floor)
    base = [int(x) for x in exact]
    # Sisa yang harus dibagikan
    remaining = total - sum(base)
    # Urutkan berdasarkan sisa terbesar (Largest Remainder), tie-break acak deterministik
    fractions = [(exact[i] - base[i], rng.random(), i) for i in range(n)]
    fractions.sort(reverse=True)
    for _, _, i in fractions:
        if remaining <= 0:
            break
        if base[i] < maxes[i]:   # jangan melebihi max komponen
            base[i] += 1
            remaining -= 1
    # Clamp ke maxes
    base = [min(base[i], maxes[i]) for i in range(n)]
    # Pastikan total tepat
    diff = total - sum(base)
    if diff > 0:
        for _, _, i in fractions:
            if diff <= 0: break
            if base[i] < maxes[i]:
                base[i] += 1; diff -= 1
    elif diff < 0:
        for _, _, i in reversed(fractions):
            if diff >= 0: break
            if base[i] > 0:
                base[i] -= 1; diff += 1
    return base


def score_group(total, key):
    """Hasilkan {col: nilai} untuk satu kelompok."""
    seed = hash(key) & 0xFFFF
    # Alokasi ke 6 komponen
    comp_vals = allocate(total, COMP_MAX, seed)
    result = {}
    for i, (cols, sub_max, _) in enumerate(COMPS):
        subs = allocate(comp_vals[i], sub_max, seed + i * 137)
        for col, val in zip(cols, subs):
            result[col] = val
    return result


# ─── Load & Fill ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX)

for kelas, kel_targets in TARGETS.items():
    if not kel_targets:
        print(f"Kelas {kelas}: skip (no data)")
        continue
    ws = wb[f"Kelas {kelas}"]
    for row in range(4, ws.max_row + 1):
        kel_raw = ws.cell(row, 4).value
        if kel_raw is None:
            break
        kel = int(kel_raw)
        if kel not in kel_targets:
            continue
        scores = score_group(kel_targets[kel], f"{kelas}_{kel}")
        for col, val in scores.items():
            ws.cell(row, col).value = val
    print(f"Kelas {kelas}: filled")

wb.save(XLSX)
print(f"\nSaved: {XLSX}")

# ─── Verifikasi cepat ────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(XLSX)
for kelas in ["A", "B", "D", "E"]:
    ws = wb2[f"Kelas {kelas}"]
    print(f"\n=== Kelas {kelas} ===")
    prev_kel = None
    for row in range(4, ws.max_row + 1):
        kel = ws.cell(row, 4).value
        if kel is None:
            break
        if kel != prev_kel:
            prev_kel = kel
            subs = [ws.cell(row, c).value or 0 for c in
                    [5,6,7, 9,10,11, 13,14,15, 17,18,19, 21,22,23, 25,26,27]]
            A = sum(subs[0:3]);  B = sum(subs[3:6])
            C = sum(subs[6:9]);  D = sum(subs[9:12])
            E = sum(subs[12:15]); F = sum(subs[15:18])
            T = A + B + C + D + E + F
            print(f"  Kel {kel:2d}: A={A:2d}/20  B={B:2d}/15  C={C:2d}/20  "
                  f"D={D:2d}/20  E={E:2d}/15  F={F:2d}/10  → TOTAL={T:3d}/100")
