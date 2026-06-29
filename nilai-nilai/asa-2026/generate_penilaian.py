"""Generate Penilaian_ASA_2026.xlsx — 5 sheet Kelas A-E"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill(c): return PatternFill("solid", fgColor=c)
def bfont(sz=10, c="000000"): return Font(bold=True, size=sz, color=c)
def nfont(sz=10, c="000000"): return Font(bold=False, size=sz, color=c)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def bdr():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)
def mbdr():
    s = Side(style="medium", color="555555")
    return Border(left=s, right=s, top=s, bottom=s)

# Warna
CH="1F3864"; CB="2E75B6"
CA="FFF2CC"; CB2="E2EFDA"; CC="EAD1DC"; CD="FCE4D6"; CE="DAEEF3"; CF="D9D9D9"
CT="FFD966"; CR="FF6B6B"; CW="FFF9C4"

# Kolom: (lbl1, lbl2, max, bg, sum_of_indices, is_nilai)
COLS=[
 ("No","",None,None,None,False),
 ("NIM","",None,None,None,False),
 ("Nama","",None,None,None,False),
 ("Kel","",None,None,None,False),
 ("A","A.1\n/10",10,CA,None,False),
 ("A","A.2\n/5",5,CA,None,False),
 ("A","A.3\n/5",5,CA,None,False),
 ("A","TOT\n/20",20,CA,(4,5,6),False),
 ("B","B.1\n/5",5,CB2,None,False),
 ("B","B.2\n/5",5,CB2,None,False),
 ("B","B.3\n/5",5,CB2,None,False),
 ("B","TOT\n/15",15,CB2,(8,9,10),False),
 ("C","C.1\n/6",6,CC,None,False),
 ("C","C.2\n/7",7,CC,None,False),
 ("C","C.3\n/7",7,CC,None,False),
 ("C","TOT\n/20",20,CC,(12,13,14),False),
 ("D","D.1\n/6",6,CD,None,False),
 ("D","D.2\n/7",7,CD,None,False),
 ("D","D.3\n/7",7,CD,None,False),
 ("D","TOT\n/20",20,CD,(16,17,18),False),
 ("E","E.1\n/5",5,CE,None,False),
 ("E","E.2\n/5",5,CE,None,False),
 ("E","E.3\n/5",5,CE,None,False),
 ("E","TOT\n/15",15,CE,(20,21,22),False),
 ("F","F.1\n/4",4,CF,None,False),
 ("F","F.2\n/3",3,CF,None,False),
 ("F","F.3\n/3",3,CF,None,False),
 ("F","TOT\n/10",10,CF,(24,25,26),False),
 ("TOTAL","/100",100,CT,(7,11,15,19,23,27),False),
 ("NILAI","Huruf",None,CT,None,True),
 ("Catatan","",None,None,None,False),
]

# Data (no, nim, nama, kel, catatan)
STUDENTS = {
"A":[
 (1,"24060124120018","Putri Elizabeth Simanjuntak",1,""),
 (2,"24060124130122","Ridho Tri Saputra",1,""),
 (3,"24060124130084","Jessica Laurencia Panjaitan",1,""),
 (4,"24060124140127","Axel Anggian Hamonangan Purba",1,""),
 (5,"24060124140194","Andiny Khaerany Suhartady",2,"Link laporan tidak dapat diakses"),
 (6,"24060124140201","Cindy Kurniawan",2,""),
 (7,"24060124130066","Silvani Salsabilla",2,""),
 (8,"24060124120007","Agung Rama Pramana Putra",2,""),
 (9,"24060124120029","Nouvella Rahma Fitrah Legarsi",3,""),
 (10,"24060124120040","Ovilia Suci Ramadhani",3,""),
 (11,"24060124130070","Biyani Andarisky Maratia",3,""),
 (12,"24060124140180","Eileen Albert Tandrio",3,""),
 (13,"24060124120048","Daniel Lamganda Tua Gultom",4,""),
 (14,"24060124130099","Dehar Zaidan Dzaki Amirullah",4,""),
 (15,"24060124130103","Dzaki Fathul'Alim Cahyo",4,""),
 (16,"24060124130108","Elang Fadila Ahmad",4,""),
 (17,"24060124120033","Shalom Kurniawan",5,"⚠ Laporan konseptual, konfirmasi 30 data"),
 (18,"24060124120009","Arsy Thariq Munawar",5,""),
 (19,"24060124120035","Yuma Hazza Yuditama",5,""),
 (20,"24060123140183","Gaza Al Ghozali Chansa",6,""),
 (21,"24060123140186","Muhammad Farhan Al Ghifari",6,""),
 (22,"24060123110046","Adriano Bawan",6,""),
 (23,"24060123130107","Muhammad Renno Baihaqi",6,""),
 (24,"24060123120033","Cacania Pasu Nalaung Siregar",7,""),
 (25,"24060123140163","Nasywa Alya Kamila",7,""),
 (26,"24060123140142","Parisya Lituhayu Chandrawati Gunawarman",7,""),
 (27,"24060124140148","Muhammad Alfaiq Rido Salafy",8,""),
 (28,"24060124140203","Mohammad Najib Fitrianto",8,""),
 (29,"24060123120012","Syafiq Abiyyu Taqi",8,""),
 (30,"24060124130069","Muhammad Fikri",9,""),
 (31,"24060124120011","Khanza Qaila",9,""),
 (32,"24060124140158","Nayla Husna",9,""),
 (33,"24060124140147","Ammar Rozan Rusyaidan",10,""),
 (34,"24060124130089","Arya Naufal Akmal",10,""),
 (35,"24060124120031","Misbachul Munir",10,""),
 (36,"24060124140154","Syahrafi Ahmad Pradika",10,""),
 (37,"24060124130114","Yustinus Hendi Setyawan",11,"Link laporan tidak dapat diakses"),
 (38,"24060124140129","Revanska Muhammad Athalla",11,""),
 (39,"24060124120046","Romualdus Yoas Wicaksono",11,""),
 (40,"24060124130076","Johan Reinhart Calvin",12,"[!] P1&P2 keduanya Jump Search, tidak ada Decrease-by-One"),
 (41,"24060124130074","Kiyoshi Akila Tira",12,""),
 (42,"24060124140198","Ilham Muhammad Raffi",12,""),
 (43,"24060124130123","Muhammad Akmal Fazli Riyadi",13,""),
 (44,"24060124120037","Muhammad Fahri",13,""),
 (45,"24060124140207","Hadrian Shandhy Yudha",13,""),
],
"B":[
 (1,"24060124130121","Fazl Nizam Priyambono",1,""),
 (2,"24060124120017","Lintang Aulia Nuraini",1,""),
 (3,"24060124120041","Nawaal Hanif Muntaz Arriye",1,""),
 (4,"24060124130094","Moses Morell Yosefan",1,""),
 (5,"24060124140179","Adam Mulya Rasyid",2,""),
 (6,"24060124140141","Muchammad Rajib Tafrichan",2,""),
 (7,"24060124140140","Muhammad Ibrahim Alghifari",2,""),
 (8,"24060124130085","Raffa Putra Nugroho",2,""),
 (9,"24060124130106","Shofwan Fikrul Huda",3,""),
 (10,"24060124120043","Syifa Aeni Mudrikah",3,""),
 (11,"24060124140146","Shafa Aqilla Zahira",3,""),
 (12,"24060124130071","Rafi Anandra Dharmawan",3,""),
 (13,"24060124120034","Diah Maulida Pratiwi",4,""),
 (14,"24060124130081","Hana Nafi'atul Haq",4,""),
 (15,"24060124120002","Birela Miadeta Purita",4,""),
 (16,"24060124120038","Alyssa Shane Kurniawan",5,""),
 (17,"24060124120050","Olivia Oktaviani",5,""),
 (18,"24060124140132","Puti Shasta Khafiyani",5,""),
 (19,"24060124140155","Revalina Salwa Aliya W. P.",5,""),
 (20,"24060124120044","Jordan Tenggara",6,""),
 (21,"24060124130078","Restu Surya",6,""),
 (22,"240601241*","Reynaldi Bertinus Hutagaol",6,"[!] NIM tidak lengkap (9 digit), wajib konfirmasi"),
 (23,"2406012414193*","Mohammad Banyuputra Eka Pramuditha",6,"[!] NIM tidak lengkap (13 digit), wajib konfirmasi"),
 (24,"24060124130077","Laurensius Brian Prayoga",7,""),
 (25,"24060124140128","Ida Bagus Ngurah Yudistira K",7,""),
 (26,"24060124120001","Rizky Saefirdaus",7,""),
 (27,"24060124140153","Iza Yunus Andhika",7,""),
 (28,"24060122140162","Zahra Nisaa Fitria Nur Afifah",8,""),
 (29,"24060122130082","Fathia Rahma",8,"Anggota ke-3 (P3) tidak terdaftar di spreadsheet"),
 (30,"24060124140152","Adhyaksa M. Banjar Nahor",9,""),
 (31,"24060124140191","Arga Yura Danendra",9,""),
 (32,"24060124140178","Raihan Lazuardi",9,""),
 (33,"24060124120025","Ganendra Satya Sindhunata",9,"Duplikat P2 dengan Arga, konfirmasi kontribusi"),
 (34,"24060123140198","Benjamin Hamonangan",10,""),
 (35,"24060123140161","Aura Cantika Nabila A.",10,""),
 (36,"24060123140128","Fakhri Ali Azadi",10,""),
 (37,"24060124140162","Galvin Shalahudin Mumtaz",11,""),
 (38,"24060124130101","Galang Bintang Ramadhan",11,""),
 (39,"24060124140185","Farhan Muhtaram",11,""),
 (40,"24060124120027","Muhammad Nauval Fadli",11,""),
 (41,"24060123140172","Cipta Fikri Wiratama",12,""),
 (42,"24060123140120","Muhammad Shafwan Raihan S",12,""),
 (43,"24060123140194","Stephen Andrew Pakpahan",12,""),
 (44,"24060124140187","Michael Stevano",13,"[!] Hanya 1 anggota terdaftar, konfirmasi kelompok"),
],
"C":[],
"D":[
 (1,"24060124140165","Dinda Isyariani",1,""),
 (2,"24060124130065","Elza Khoirisma Carrynda",1,""),
 (3,"24060124130095","Zulfa Nabilah",1,""),
 (4,"24060121120026","M. Ghani Aryasuta",1,"Angkatan 2021, konfirmasi status aktif"),
 (5,"24060124140196","Saburo Rafqi Hidayat",2,""),
 (6,"24060124120030","An Al Rivaldi",2,""),
 (7,"24060124130054","Raffie Aditya Akbar",2,""),
 (8,"24060124130092","Syuraih Umar Khottob",2,""),
 (9,"24060124140170","Naufal Rayan Attallah",3,""),
 (10,"24060124130056","Muhammad Kamal Hamzah",3,""),
 (11,"24060124140126","Rafa Azlan",3,""),
 (12,"24060124130075","Naufal Dwi Yusmawan",3,""),
 (13,"24060124120036","Marchella Arkhina Ratunesia",4,""),
 (14,"24060124120004","Kayla Febrina Laura Ayu",4,""),
 (15,"24060124140150","Nashwa Aldebaran",4,""),
 (16,"24060123140130","Ratu Sekar Ayu Nisa",5,""),
 (17,"24060123140131","Mutiara Ayu Pramono",5,""),
 (18,"24060123140174","Duta Adi Pamungkas",5,""),
 (19,"24060124140133","Muhammad Hafidh Al-Ghifari",6,""),
 (20,"24060124140164","Muhammad Hafidh Zufar Dewantara",6,""),
 (21,"-","Muhammad Lutfi Febriansyah",6,"[!] NIM tidak tercatat, wajib konfirmasi"),
 (22,"-","Adefritz Einar Sinaga",6,"[!] NIM tidak tercatat, wajib konfirmasi"),
 (23,"24060124130118","Dewangga Ramadhan Halim",7,""),
 (24,"24060124140145","Ferdy Prasetya Putra",7,""),
 (25,"24060124120051","Caesar Ferdiana Suwandi",7,""),
 (26,"24060124140171","Ghatfan Muhammad Atiwiar",7,"Ditandai P di spreadsheet, konfirmasi partisipasi"),
 (27,"24060123140184","Muhammad Kievlan Hakim",8,""),
 (28,"24060123140171","Muhammad Iman Sasongko",8,""),
 (29,"24060123130067","Nindya Kirana",8,""),
 (30,"24060123140045","Dewi Larasati Mumpuni",8,""),
 (31,"24060124130107","Muhammad Firdaus Argifari",9,""),
 (32,"24060124140160","Basil Ayman Hariadi",9,""),
 (33,"24060124140130","Crystiano Bayu Satya Alves",9,""),
 (34,"24060124140149","Devano Trestanto",10,""),
 (35,"24060124120008","Amelia Aristianti",10,""),
 (36,"24060124140190","Mohammad Andhika Ramadhan",10,""),
 (37,"24060124140174","Novelya Cherina",10,""),
 (38,"24060124130119","Hasta Putra Wildantara",11,""),
 (39,"24060124140183","Hanif Ihsanul Huda",11,""),
 (40,"24060124120039","Izzatu Khoirul Fata",11,""),
 (41,"24060124140202","Ikrar Maheswara Rabbani Wibowo",12,""),
 (42,"24060124140138","Menza Isaiah Tampubolon",12,""),
 (43,"24060124130072","Ali Maskan Ferry Purwanto",12,""),
 (44,"24060124140161","Azka Wayasy Al Hafizh",12,""),
],
"E":[
 (1,"24060123140190","Tegar Caesara",1,""),
 (2,"24060123140132","Zaky Musyaffa",1,""),
 (3,"24060123120043","Sophie Venecia M",1,""),
 (4,"24060124120015","Wahyu Eko Setyo P",1,""),
 (5,"24060124130113","Joshua Satria Kusuma",2,""),
 (6,"24060124140131","Harits Permana",2,""),
 (7,"24060124130097","Haikal Imam Ridha",2,""),
 (8,"24060124120047","Husein Avicenna",2,""),
 (9,"24060124120012","Felicia Evelina",3,""),
 (10,"24060124140188","Claudia Meitania Putri",3,""),
 (11,"24060124130053","Anintya Abhi Wiryateja",3,""),
 (12,"24060124120006","Aditya Sultonul Ulya",3,""),
 (13,"24060124130115","Rafif Setya Imaduddin",4,""),
 (14,"24060124140124","Aqiatillah Rezi Zhafran",4,""),
 (15,"24060124120020","Raffi Arditama",4,""),
 (16,"24060124140166","Muhammad Farhan Abdul Azis",4,""),
 (17,"24060124120026","Gregorius Septiano Ariadi",5,""),
 (18,"24060124120010","Dhimas Reza Nafi Wahyudi",5,""),
 (19,"24060124140169","Fadhil Yaafi Widodo",5,""),
 (20,"24060124140137","Farhan Dwiyan Akbar",5,""),
 (21,"24060124130064","Anggita Kirana P.",6,""),
 (22,"24060124140125","Varissa Nabila Kifli",6,""),
 (23,"24060124140204","Adelia Clearesta",6,""),
 (24,"24060124140134","Wahyu Aji Gumelar Tri Nugroho",6,""),
 (25,"24060124120019","Nadia Azura Nurhaniya",7,""),
 (26,"24060124120013","Muhamad Kemal Faza",7,""),
 (27,"24060124110142","Muchamad Yuda Tri Ananda",7,"Kode 110, konfirmasi program studi"),
 (28,"24060124140200","Muhammad Zaidaan Ardiyansyah",7,""),
 (29,"24060124130086","Aron Sorimuda Johanes Pasaribu",8,""),
 (30,"24060124130055","Daffa Maulana Alfianto",8,""),
 (31,"24060124140163","Djuan Setyo Jati",8,""),
 (32,"24060124130067","Levi Ramot Siahaan",8,""),
 (33,"24060123130099","Abyasa Saifaji",9,""),
 (34,"24060123140156","Alfonso Michael N. H",9,""),
 (35,"24060122140178","Khairindra Eka Putra",9,"Angkatan 2022, konfirmasi status aktif"),
 (36,"24060122120026","Tiara Putri W.",9,"Angkatan 2022, konfirmasi status aktif"),
 (37,"24060124120016","Quinta Aurabiansyah",10,""),
 (38,"24060124140139","Muhammad Fauzan Akbar",10,""),
 (39,"24060124130061","Rahmat Argyandha Aminuddin",10,""),
 (40,"24060124140177","Rayhan Gerard Darmawan",10,""),
 (41,"24060123140191","Muhammad Iqbal Haqiqi",11,""),
 (42,"24060123130055","Novanza Edgar Wibowo",11,""),
 (43,"24060123140063","Bagas Nur Ardianto",11,""),
],
}

# Warna alternasi per kelompok (genap/ganjil)
GRP_COLORS = ["FFFFFF","EBF3FB"]  # putih / biru sangat muda

def nilai_formula(total_col_letter, row):
    tc = total_col_letter
    r  = row
    return (f'=IF({tc}{r}="","",IF({tc}{r}>=85,"A",IF({tc}{r}>=75,"AB",'
            f'IF({tc}{r}>=65,"B",IF({tc}{r}>=55,"BC",IF({tc}{r}>=45,"C",'
            f'IF({tc}{r}>=35,"CD",IF({tc}{r}>=25,"D","E"))))))))')

def build_sheet(wb, kelas, students):
    ws = wb.create_sheet(title=f"Kelas {kelas}")
    ncols = len(COLS)

    # ── Baris 1: Judul ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1)
    t.value = f"LEMBAR PENILAIAN TUGAS KELOMPOK — ASA KELAS {kelas} 2026"
    t.font = bfont(13, "FFFFFF")
    t.fill = fill(CH)
    t.alignment = ctr()
    t.border = bdr()
    ws.row_dimensions[1].height = 28

    # ── Baris 2-3: Header komponen (merge per grup) ─────────────────────────
    # Baris 2: label utama (A,B,C,D,E,F,TOTAL,NILAI,Catatan), identitas
    # Baris 3: sub-label (A.1, A.2, ...)
    # Cari posisi awal tiap komponen untuk merge
    comp_start = {}
    for ci, (l1,l2,mx,bg,sumof,isval) in enumerate(COLS):
        if l1 not in comp_start:
            comp_start[l1] = ci
    comp_ranges = {}
    cur_l1 = None; cur_start = 0
    for ci, (l1,l2,mx,bg,sumof,isval) in enumerate(COLS):
        if l1 != cur_l1:
            if cur_l1 is not None:
                comp_ranges[cur_l1] = (cur_start, ci-1)
            cur_l1 = l1; cur_start = ci
    if cur_l1 is not None:
        comp_ranges[cur_l1] = (cur_start, len(COLS)-1)

    # Row 2: merge labels
    for l1,(s,e) in comp_ranges.items():
        col_s = s+1; col_e = e+1
        if col_s == col_e:
            c = ws.cell(2, col_s)
            c.value = l1
        else:
            ws.merge_cells(start_row=2,start_column=col_s,end_row=2,end_column=col_e)
            c = ws.cell(2, col_s)
            c.value = l1
        bg = COLS[s][3]
        c.font = bfont(10, "FFFFFF" if l1 in ("TOTAL","NILAI","A","B","C","D","E","F") else "000000")
        if l1 == "TOTAL" or l1 == "NILAI": bg_use = CT
        elif bg: bg_use = bg
        else: bg_use = CB
        c.fill = fill(bg_use)
        c.alignment = ctr()
        c.border = bdr()
    ws.row_dimensions[2].height = 22

    # Row 3: sub-labels
    for ci,(l1,l2,mx,bg,sumof,isval) in enumerate(COLS):
        c = ws.cell(3, ci+1)
        c.value = l2 if l2 else l1
        bg_use = bg if bg else ("FFFFFF" if ci < 4 else CB)
        if l1 in ("TOTAL","NILAI"): bg_use = CT
        c.fill = fill(bg_use)
        c.font = bfont(9)
        c.alignment = ctr()
        c.border = bdr()
    ws.row_dimensions[3].height = 36

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "E4"

    # ── Data rows ────────────────────────────────────────────────────────────
    if not students:
        ws.cell(4,1).value = "Data belum tersedia — lihat penilaian-asa-c.md"
        ws.cell(4,1).font = bfont(10, "FF0000")
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
        _set_col_widths(ws)
        return

    prev_kel = None; grp_idx = -1
    for row_idx, (no,nim,nama,kel,cat) in enumerate(students):
        r = row_idx + 4  # data start row 4
        if kel != prev_kel:
            grp_idx += 1
            prev_kel = kel
        row_bg = GRP_COLORS[grp_idx % 2]

        for ci,(l1,l2,mx,bg,sumof,isval) in enumerate(COLS):
            c = ws.cell(r, ci+1)
            # Identitas
            if ci == 0: c.value = no
            elif ci == 1: c.value = str(nim)
            elif ci == 2: c.value = nama
            elif ci == 3: c.value = kel
            elif ci == len(COLS)-1:   # Catatan
                c.value = cat
                if cat and ("[!" in cat or "Angkatan 20" in cat or "konfirmasi" in cat.lower()):
                    c.font = nfont(9, "C00000")
                    c.fill = fill("FFF2CC")
                else:
                    c.font = nfont(9)
                    c.fill = fill(row_bg)
                c.alignment = lft()
                c.border = bdr()
                continue
            elif isval:   # Nilai huruf
                total_col = get_column_letter(len(COLS)-2)  # kolom TOTAL
                c.value = nilai_formula(total_col, r)
                c.font = bfont(10)
            elif sumof is not None:   # formula sum
                parts = "+".join(f"{get_column_letter(si+1)}{r}" for si in sumof)
                counta_args = ",".join(f"{get_column_letter(si+1)}{r}" for si in sumof)
                c.value = f"=IF(COUNTA({counta_args})=0,\"\",{parts})"
                c.font = bfont(10)
            # Styling
            if ci < 4:
                c.fill = fill(row_bg)
                c.font = nfont(10) if ci != 2 else nfont(10)
                c.alignment = ctr() if ci != 2 else lft()
            else:
                cell_bg = bg if bg else row_bg
                if l1 in ("TOTAL","NILAI"): cell_bg = CT
                c.fill = fill(cell_bg)
                c.alignment = ctr()
                if sumof or isval:
                    c.font = bfont(10)
            c.border = bdr()
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws)

def _set_col_widths(ws):
    widths = [5,18,38,5,6,5,5,7,5,5,5,7,5,6,6,7,5,6,6,7,5,5,5,7,5,4,4,7,9,7,32]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Main ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # hapus sheet default

for kelas in ["A","B","C","D","E"]:
    build_sheet(wb, kelas, STUDENTS[kelas])

out = "/projects/sandbox/sonet-product/nilai-nilai/asa-2026/Penilaian_ASA_2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
