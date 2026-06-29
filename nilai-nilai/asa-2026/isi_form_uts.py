"""
Pindahkan nilai Tugas (Decrease and Conquer / Tugas Sebelum UTS) ke kolom
"Nilai UTS" pada form SIAP MIK1624404 (ASA Teori) untuk kelas A-E.

- Form target : SIAP_TemplateNilai_MIK1624404_2025_2_{A-E}_gabungan.xlsx
- Sheet       : "Worksheet"  (header baris 7, data mulai baris 8)
- Kolom NIM   : A   |  Kolom Nama : B  |  Kolom UTS (target) : I
- Pencocokan  : berdasarkan NIM.

Catatan: PAIK6601 (ASA Praktikum) TIDAK diisi karena merupakan komponen
mata kuliah berbeda dengan roster praktikum tersendiri.
"""
import openpyxl

DIR = "/projects/sandbox/sonet-product/nilai-nilai/asa-2026"

# Target total per kelompok (sumber: fill_nilai.py / PR #107)
TARGETS = {
  "A": {1:82,2:64,3:87,4:81,5:67,6:89,7:78,8:84,9:80,10:86,11:65,12:69,13:83},
  "B": {1:85,2:81,3:88,4:78,5:83,6:73,7:84,8:71,9:77,10:82,11:86,12:91,13:68},
  "C": {},
  "D": {1:79,2:82,3:85,4:78,5:74,6:71,7:83,8:87,9:76,10:81,11:89,12:77},
  "E": {1:84,2:80,3:87,4:83,5:78,6:86,7:74,8:81,9:76,10:91,11:79},
}

# Build NIM -> nilai dari Penilaian_ASA_2026.xlsx (kolom B=NIM, D=Kelompok)
wb_src = openpyxl.load_workbook(f"{DIR}/Penilaian_ASA_2026.xlsx")
nim_grade = {}
for kelas in ["A", "B", "C", "D", "E"]:
    ws = wb_src[f"Kelas {kelas}"]
    for r in range(4, ws.max_row + 1):
        nim = ws.cell(r, 2).value
        kel = ws.cell(r, 4).value
        if not nim or kel is None:
            continue
        tot = TARGETS.get(kelas, {}).get(int(kel))
        if tot is not None:
            # buang tanda "*" pada NIM tidak lengkap
            nim_grade[str(nim).strip().rstrip("*")] = tot

# Resolusi NIM lengkap untuk mahasiswa yang NIM-nya tidak lengkap / kosong
# di spreadsheet tugas, tetapi ada di form resmi (dicocokkan via nama).
MANUAL = {
    "24060124140157": 73,  # Reynaldi Bertinus Hutagaol  (B, Kel 6)
    "24060124140193": 73,  # Mohammad Banyuputra Eka Pramudhita (B, Kel 6)
    "24060124140197": 71,  # Muhammad Lutfi Febriansyah  (D, Kel 6)
    "24060124140186": 71,  # Adefritz Einar Sinaga       (D, Kel 6)
}
nim_grade.update(MANUAL)

UTS_COL = 9   # kolom I = Nilai UTS
NIM_COL = 1   # kolom A
NAMA_COL = 2  # kolom B
DATA_START = 8

print(f"Total NIM dengan nilai: {len(nim_grade)}\n")

grand_filled = 0
grand_unmatched = []
for cls in ["A", "B", "C", "D", "E"]:
    fn = f"{DIR}/SIAP_TemplateNilai_MIK1624404_2025_2_{cls}_gabungan.xlsx"
    wb = openpyxl.load_workbook(fn)
    ws = wb["Worksheet"]
    filled = 0
    unmatched = []
    for r in range(DATA_START, ws.max_row + 1):
        nim = ws.cell(r, NIM_COL).value
        if not nim:
            continue
        nim = str(nim).strip()
        if nim in nim_grade:
            ws.cell(r, UTS_COL).value = nim_grade[nim]
            filled += 1
        else:
            unmatched.append((nim, ws.cell(r, NAMA_COL).value))
    wb.save(fn)
    grand_filled += filled
    for u in unmatched:
        grand_unmatched.append((cls, *u))
    print(f"MIK1624404_{cls}: terisi {filled}, belum {len(unmatched)}")
    for nim, nama in unmatched:
        print(f"    [kosong] {nim}  {nama}")

print(f"\nTOTAL kolom UTS terisi: {grand_filled}")
print(f"TOTAL belum terisi    : {len(grand_unmatched)}")
