#!/usr/bin/env python3
"""
GTI 2026 — Penilaian Otomatis via YouTube Description Analysis
Fetch deskripsi video → analisis kata kunci → skor rubrik → Excel
"""
import requests, re, json, time, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============================================================
# 1. VIDEO METADATA
# ============================================================
ALL_VIDEOS = {
    'A3':'Im8KWkQt2oA','A4':'_jsEwoba0yQ','A5':'snYwaR38XQ4',
    'A6':'Z59d6nk4PAs','A7':'oQcXjDnyGJU','A8':'f-mvyIQzta4',
    'A9':'donL48coYQE','A10':'NOa6aFGw6cs','A11':'PEcKOSPsxIc','A13':'FltKiLTQOOE',
    'B1':'8CHXPfwOpLE','B2':'ke7Sop09qNc','B3':'kPP7qwyveFU',
    'B4':'w0AvKSJOdfg','B5':'X7-EdNZdfr8','B6':'OxwGDezCivo',
    'B7':'rWW6hoEYtHc','B8':'EqYGuwdQ4b8','B9':'rOyCfo1IjeY',
    'B10':'VV_jRwETrEY','B11':'Q-S8BUXXiSE',
    'C3':'3wSXNM-3HzA','C4':'dBWuYekfPr0','C5':'_HqhHM9IX80',
    'C6':'KcBhLfcCoas','C7':'HQRgEdBYV2k','C8':'OrlicuJV3tQ',
    'C9':'ewmR8Pro5bg','C11':'IpFz4pn92Ow','C12':'rEJk762uNYw',
    'D1':'u5T9oImjFh8','D2':'9r7-SSIk6kQ','D3':'zBAv8qhIzN4',
    'D4':'3ojOKIl4RUs','D5':'18sq-LZdQtc','D6':'EYOkbbIlPjc',
    'D7':'LGzCdTaglz0','D8':'kUmYg1qulW4','D9':'6FlXx3K7jj8','D10':'5uoeRSHvfQI',
    'E2':'uxiIlfewLVs','E4':'8rbqDYMV0Bo','E9':'IVyrxIBSRzE',
}

# ============================================================
# 2. FETCH YOUTUBE DATA
# ============================================================
def build_session():
    s = requests.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0',
        'Accept-Language': 'id,en-US;q=0.7,en;q=0.3',
        'Accept-Encoding': 'identity',
    })
    s.get('https://www.youtube.com/', timeout=8)
    return s

def fetch_yt(session, vid_id):
    try:
        r = session.get(f'https://www.youtube.com/watch?v={vid_id}', timeout=15)
        text = r.text
        # Description
        desc = ''
        m = re.search(r'"attributedDescriptionBodyText":\{"content":"(.*?)"(?:,"styleRuns"|,"commandRuns"|\})', text, re.DOTALL)
        if m: desc = m.group(1).replace('\\n','\n').replace('\\"','"').replace('\\\\','\\')
        if not desc:
            m2 = re.search(r'"shortDescription":"((?:[^"\\]|\\.)*?)","isCrawlable"', text)
            if m2: desc = m2.group(1).replace('\\n','\n').replace('\\"','"')
        # Title via oEmbed (more reliable)
        try:
            ro = requests.get(f'https://www.youtube.com/oembed?url=https://www.youtube.com/watch?v={vid_id}&format=json', timeout=6)
            title = ro.json().get('title','')
        except:
            title = ''
        return {'title': title, 'desc': desc, 'ok': bool(desc or title)}
    except Exception as e:
        return {'title': '', 'desc': '', 'ok': False, 'err': str(e)}

def fetch_all():
    print("Fetching YouTube data...")
    session = build_session()
    results = {}
    items = list(ALL_VIDEOS.items())
    with ThreadPoolExecutor(max_workers=5) as ex:
        futs = {ex.submit(fetch_yt, session, vid): key for key, vid in items}
        for fut in as_completed(futs):
            key = futs[fut]
            data = fut.result()
            data['vid'] = ALL_VIDEOS[key]
            results[key] = data
            status = "✓" if data['ok'] else "✗"
            print(f"  {status} {key}: {data['title'][:55]}")
    return results


# ============================================================
# 3. RUBRIK SCORING ENGINE
# ============================================================

RUBRIK_KEYWORDS = {
    'B': {  # Representasi Objek 3D
        'strong': ['vertex','edge','face','mesh','polygon','primitif','primitive',
                   'kubus','bola','kerucut','silinder','cylinder','cube','sphere','cone',
                   'bangun 3d','objek 3d','object 3d','geometri 3d','geometry 3d',
                   'vertices','faces','kotak','piramida','plane'],
        'medium': ['3d','tiga dimensi','three dimensional','bentuk','shape','model','objek','object'],
        'max': 10
    },
    'C': {  # Proyeksi Orthographic
        'strong': ['orthographic','ortho','ortografik','tampak atas','tampak depan','tampak samping',
                   'top view','front view','side view','parallel projection','proyeksi sejajar',
                   'view ortho','parallel','top-down'],
        'medium': ['proyeksi','projection','tampak','view','pandangan'],
        'max': 10
    },
    'D': {  # Proyeksi Perspektif
        'strong': ['perspektif','perspective','1 point','2 point','3 point','one point','two point',
                   'vanishing','titik hilang','frustum','fov','field of view','proyeksi perspektif',
                   '1-point','2-point'],
        'medium': ['kamera perspektif','depth','kedalaman','jauh dekat','fov','sudut pandang'],
        'max': 10
    },
    'E': {  # Geometri / Transformasi
        'strong': ['translasi','rotasi','skala','transformasi','translation','rotation','scale',
                   'transform','matrix transformasi','matriks','bergerak','berputar','scale transform',
                   'glrotatef','gltranslatef','glscalef','translate','rotate'],
        'medium': ['gerak','pergerakan','movement','posisi','position','bergerak','animasi','animation'],
        'max': 8
    },
    'F': {  # Kamera
        'strong': ['kamera','camera','fov','field of view','viewport','clipping','frustum',
                   'lookat','look at','up vector','eye position','near plane','far plane',
                   'gluperspective','glulookat','view matrix','projection matrix'],
        'medium': ['sudut kamera','posisi kamera','camera position','view','tampilan','pandangan kamera'],
        'max': 8
    },
    'G': {  # Cahaya
        'strong': ['cahaya','lighting','ambient','diffuse','specular','shininess',
                   'directional light','point light','spot light','pencahayaan','iluminasi',
                   'gllightfv','gl_light','gl_ambient','gl_diffuse','cahaya ambien','cahaya difus',
                   'phong lighting','blinn','illumination'],
        'medium': ['terang','gelap','bayangan','shadow','bright','light','sumber cahaya','lampu'],
        'max': 8
    },
    'H': {  # Karakteristik Permukaan
        'strong': ['material','texture','tekstur','roughness','metallic','permukaan',
                   'surface','warna material','material property','glmaterialfv','gl_material',
                   'reflection','refleksi','specular map','normal map','bump map',
                   'emissive','albedo','uv mapping','bahan'],
        'medium': ['warna','color','colour','tampilan permukaan','koefisien','coefficient'],
        'max': 8
    },
    'I': {  # Algoritma Rendering
        'strong': ['rendering','rasterization','rasterisasi','pipeline','wireframe','scanline',
                   'raycast','ray cast','opengl','glut','fragment shader','vertex shader',
                   'depth buffer','z-buffer','hidden surface','painter','render pipeline',
                   'backface culling','culling'],
        'medium': ['render','gambar','display','tampil','frame','buffer','grafis'],
        'max': 8
    },
    'J': {  # Shading
        'strong': ['shading','flat shading','gouraud','phong','normal map','shader',
                   'warna permukaan','lighting model','smooth shading','constant shading',
                   'interpolasi warna','color interpolation','gl_smooth','gl_flat',
                   'shade model','glshademodel'],
        'medium': ['pencahayaan permukaan','gradien warna','warna cahaya','normal','surface color'],
        'max': 10
    },
}

def score_component(text_lower, comp_key, max_score):
    kw = RUBRIK_KEYWORDS[comp_key]
    strong_hits = sum(1 for k in kw['strong'] if k in text_lower)
    medium_hits = sum(1 for k in kw['medium'] if k in text_lower)

    if strong_hits >= 4:
        base = 0.92
    elif strong_hits >= 2:
        base = 0.80
    elif strong_hits == 1:
        base = 0.67
    elif medium_hits >= 2:
        base = 0.57
    elif medium_hits == 1:
        base = 0.47
    else:
        base = 0.38  # video ada, konten visual tidak bisa dinilai

    return round(max_score * base)

def score_A(title, desc):
    """Kualitas Video BTS (max 20)"""
    score = 0
    lower = (title + ' ' + desc).lower()
    # Cek format judul: ada BTS dan nama panggilan dalam kurung
    if re.search(r'bts', title.lower()): score += 5
    elif re.search(r'behind the scene', title.lower()): score += 4
    else: score += 2
    # Deskripsi ada NIM (pola 240601...)
    nim_count = len(re.findall(r'240\d{9}', desc))
    if nim_count >= 3: score += 5
    elif nim_count >= 2: score += 4
    elif nim_count >= 1: score += 3
    else: score += 1
    # Deskripsi ada penjelasan teknis
    tech_words = ['teknik','technique','implementasi','implement','rendering','proyeksi',
                  'projection','kamera','camera','cahaya','lighting','material','texture',
                  'opengl','glut','3d','transformasi']
    tech_hits = sum(1 for w in tech_words if w in lower)
    if tech_hits >= 5: score += 5
    elif tech_hits >= 3: score += 4
    elif tech_hits >= 1: score += 3
    else: score += 1
    # Video accessible
    if desc: score += 5
    else: score += 0
    return min(score, 20)

def grade_from_total(total):
    if total >= 85: return 'A'
    elif total >= 75: return 'AB'
    elif total >= 65: return 'B'
    elif total >= 55: return 'BC'
    elif total >= 45: return 'C'
    elif total >= 35: return 'D'
    else: return 'E'

def score_video(key, title, desc):
    if not title and not desc:
        return None  # video tidak bisa diakses
    text = (title + '\n' + desc).lower()
    scores = {
        'A': score_A(title, desc),
        'B': score_component(text, 'B', 10),
        'C': score_component(text, 'C', 10),
        'D': score_component(text, 'D', 10),
        'E': score_component(text, 'E', 8),
        'F': score_component(text, 'F', 8),
        'G': score_component(text, 'G', 8),
        'H': score_component(text, 'H', 8),
        'I': score_component(text, 'I', 8),
        'J': score_component(text, 'J', 10),
    }
    total = sum(scores.values())
    scores['total'] = total
    scores['grade'] = grade_from_total(total)
    return scores


# ============================================================
# 4. STUDENT DATA (nama, NIM, proyek) per kelas
# ============================================================
# Format: (kelompok, nama, nim, nama_proyek_default)
# NIM '-' = belum ada data
# nama_proyek diisi dari spreadsheet; link diisi dari video fetch

DATA_GTI_A = [
    (1, '-','-','-'),
    (2, '-','-','-'),
    (3,'Johan Reinhart Calvin','24060124130076','-'),
    (3,'Ilham Muhammad Raffi','24060124140198','-'),
    (3,'Romualdus Yoas Wicaksono','24060124120046','-'),
    (3,'Muhammad Akmal Fazli Riyadi','24060124130123','-'),
    (4,'Muhammad Fikri','24060124130069','-'),
    (4,'Khanza Qaila','24060124120011','-'),
    (4,'Nayla Husna','24060124140158','-'),
    (5,'Shalom Kurniawan','24060124120033','-'),
    (5,'Arsy Thariq Munawar','24060124120009','-'),
    (5,'Yuma Hazza Yuditama','24060124120035','-'),
    (5,'Fernanda Galih Saputra','24060121140176','-'),
    (6,'Silvani Salsabilla','24060124130066','-'),
    (6,'Eileen Albert Tandrio','24060124140180','-'),
    (6,'Jessica Laurencia Panjaitan','24060124140084','-'),
    (7,'Muhammad Fahri','24060124120037','-'),
    (7,'Nouvella Rahma Fitrah Legarsi','24060124120029','-'),
    (7,'Ovilia Suci Ramadhani','24060124120040','-'),
    (7,'Biyani Andarisky Maratia','24060124130070','-'),
    (8,'Agung Rama Pramana Putra','24060124120007','-'),
    (8,'Andiny Khaerany Suhartady','24060124140194','-'),
    (8,'Hadrian Shandhy Yudha','24060124140207','-'),
    (8,'Mohammad Najib Fitrianto','24060124140203','-'),
    (9,'Elang Fadila Ahmad','24060124130108','-'),
    (9,'Dehar Zaidan Dzaki Amirullah','24060124130099','-'),
    (9,'Ridho Tri Saputra','24060124130122','-'),
    (9,'Axel Anggian Hamonangan Purba','24060124140127','-'),
    (10,'Ammar Rozan Rusyaidan','24060124140147','-'),
    (10,'Arya Naufal Akmal','24060124130089','-'),
    (10,'Misbachul Munir','24060124120031','-'),
    (11,"Dzaki Fathul'Alim Cahyo",'24060124130103','-'),
    (11,'Daniel Lamganda Tua Gultom','24060124120048','-'),
    (11,'Putri Elizabeth Simanjuntak','24060124120018','-'),
    (12,'-','-','-'),
    (13,'Revanska Athallah Muhammad','24060124140129','-'),
    (13,'Kiyoshi Akila Tira','24060124130074','-'),
    (13,'Yustinus Hendi Setyawan','24060124130114','-'),
]

DATA_GTI_B = [
    (1,'Muchammad Rajib Tafrichan','24060124140141','Car Drifting Animation'),
    (1,'Fazl Nizam Priyambodho','24060124130121',''),
    (1,'Muhammad Ibrahim Alghifari','24060124140140',''),
    (1,'Adam Mulya Rasyid','24060124140179',''),
    (2,'Moses Morell Yosefan','24060124130094','Furiosa Riding'),
    (2,'Lintang Aulia Nuraini','24060124120017',''),
    (2,'Alyssa Shane Kurniawan','24060124120038',''),
    (2,'Olivia Oktaviani','24060124120050',''),
    (3,'Shafa Aqilla Zahira','24060124140146','Look Around'),
    (3,'Rafi Anandra Dharmawan','24060124130071',''),
    (3,'Shofwan Fikrul Huda','24060124130106',''),
    (4,'Birela Miadeta Purita','24060124120002','3D TilesRush'),
    (4,'Diah Maulida Pratiwi','24060124120034',''),
    (4,"Hana Nafi'atul Haq",'24060124130081',''),
    (4,'Arga Yura Danendra','24060124140191',''),
    (5,'Jordan Tenggara','24060124120044','Hamsterball Rolling Game'),
    (5,'Reynaldi Bertinus Hutagaol','24060124140157',''),
    (5,'Mohammad Banyuputra Eka Pramuditha','24060124140193',''),
    (5,'Ida Bagus Ngurah Yudistira Kemenuh','24060124140128',''),
    (6,'Syifa Aeni Mudrikah','24060124120043','Love Simulator Game'),
    (6,'Galang Bintang Ramadhan','24060124130101',''),
    (6,'Puti Shasta Khafiyani','24060124140132',''),
    (6,'Revalina Salwa Aliya Wicaksono Prabowo','24060124140155',''),
    (7,'Ananda Bagus Tri Utomo','24060122130091','Truck Simulator'),
    (7,'Muhamad Hafidz Zulfikar','24060122140141',''),
    (7,'Nabil Razaki Herman','24060122140147',''),
    (7,'Rafi Deandra','24060122140122',''),
    (8,'Raffa Putra Nugroho','24060124130085','3D Train Simulator'),
    (8,'Laurensius Brian Prayoga','24060124130077',''),
    (8,'Iza Yunus Andhika','24060124140153',''),
    (9,'Nawaal Hanif Mumtaz Arriye','24060124120041','EZ Flappy Bird Game'),
    (9,'Galvin Shalahudin Mumtaz','24060124140162',''),
    (9,'Rizky Saefirdaus','24060124120001',''),
    (9,'Muhammad Nauval Fadli','24060124120027',''),
    (10,'Adhyaksa Margandatua Banjar Nahor','24060124140152','Quiz Maze Escape Game'),
    (10,'Raaihan Lazuardi','24060124140178',''),
    (10,'Farhan Muhtarram','24060124140185',''),
    (10,'Ganendra Satya Sindhunata','24060124120025',''),
    (11,'Michael Stevano','24060124140187','Tower Stacker Game'),
]


DATA_GTI_C = [
    (1,'-','-','-'),(2,'-','-','-'),
    (3,'Ruth Septriana Sipangkar','24060124120024','-'),
    (3,'Sarifa Nuha Ardanti Jusmar','24060124130082','-'),
    (3,'Syafira Azka Ramadhani','24060124130088','-'),
    (3,'Yasmina Syahidah','24060124130116','-'),
    (4,'Akbar Mukti Wibowo','24060124130063','-'),
    (4,'Maulana Ghazzam Adil Al Faiq','24060124130083','-'),
    (4,'Muhammad Izzat Fauzan Putra Arya','24060124130096','-'),
    (4,'Muhammad Rofad Hamdani','24060124130117','-'),
    (5,'Annis Fakhiroh Akbar','24060124130110','-'),
    (5,'Binar Ridha Wiritanaya','24060124140143','-'),
    (5,'Nabila Kayla Rafa','24060124120022','-'),
    (6,'Farras Hilmy Zaidan','24060124120003','-'),
    (6,'Imam Alfarezzel','24060124120028','-'),
    (6,'Marco Falias Pangkado','24060124130112','-'),
    (6,"Haydar Rafi' Sultansyah",'24060124120023','-'),
    (7,'Mohammad Sulthon Ariefin','24060124130104','-'),
    (7,'Haikal Rafli Sembiring','24060124130079','-'),
    (7,'Rio Setiawan Hastanu Putra','24060124130068','-'),
    (7,'Naufal Akbar Nugroho','24060124130057','-'),
    (8,'Azka Aqylla Maulana','24060124140195','-'),
    (8,'Akmal Kafli Anan','24060124120042','-'),
    (8,'Adel Rayyan Hakim','24060124140173','-'),
    (8,'Agil Yudis Wibawa','24060124120045','-'),
    (9,'Husni Ulyaa Khanifah','24060124120021','-'),
    (9,'Dian Berlian Hutasoit','24060124120005','-'),
    (9,'Christianna Olivia Juniarti M','24060124140168','-'),
    (9,'Dian Aulya Dewiyani','24060124130059','-'),
    (10,'Aswalila Adha Putri Telaumbanua','24060124120014','-'),
    (10,'Alodia Evelyn Pratikno','24060124130087','-'),
    (10,'Arini Latifatul Qalbiah','24060124140136','-'),
    (10,'Aprillia Abel Cleodora','24060124140176','-'),
    (11,'Muhammad Abhista Pratama Sava','24060124130058','-'),
    (11,'Wipin Saputra Poh','24060124130080','-'),
    (12,'Aufaarel Nabiil Aryadh Mecca','24060124140206','-'),
    (12,'Muhammad Zaidan Alfarizi','24060124130102','-'),
    (12,'Shifa Buja Jauza','24060124140182','-'),
]

# GTI D — NIM/nama diambil dari deskripsi video saat runtime
DATA_GTI_D_PLACEHOLDER = [(i, '-', '-', '-') for i in range(1, 11)]

DATA_GTI_E = [
    (1,'-','-','-'),
    (2,'Varissa Nabila Kifli','24060124140125','Game Maze 3D'),
    (2,'Wahyu Aji Gumelar Tri Nugroho','24060124140134',''),
    (2,'Claudia Meitania Putri','24060124140188',''),
    (2,'Adelia Clearesta','24060124140204',''),
    (3,'Farhan Dwiyan Akbar','24060124140137','Game Runner'),
    (3,'Haikal Imam Ridha','24060124130097',''),
    (3,'Harits Permana','24060124140131',''),
    (3,'Joshua Satria Kusuma','24060124130113',''),
    (4,'Aron Sorimuda Johanes Pasaribu','24060124130086','Sistem Tata Surya'),
    (4,'Husein Avicenna','24060124120047',''),
    (4,'Rahmat Argyandha Aminuddin','24060124130061',''),
    (5,'Gregorius Septiano Ariadi','24060124120026','Baldis Game'),
    (5,'Dhimas Reza Nafi Wahyudi','24060124120010',''),
    (5,'Djuan Setyo Jati','24060124140163',''),
    (5,'Fadhil Yaafi Widodo','24060124140169',''),
    (6,'Anggita Kirana Puspa','24060124130064','ZombieVerse'),
    (6,'Felicia Evelina','24060124120012',''),
    (6,'Raffi Arditama','24060124120020',''),
    (6,'Rafif Setya Imaduddin','24060124130115',''),
    (7,'Nadia Azura Nurhaniya','24060124120019','Backroom Escape'),
    (7,'Muchamad Yuda Tri Ananda','24060124110142',''),
    (7,'Aditya Sultonul Ulya','24060124120006',''),
    (7,'Rayhan Gerard Darmawan','24060124140177',''),
    (8,'Aqiatillah Rezi Zhafran','24060124140124','FreakyMaze'),
    (8,'Muhammad Farhan Abdul Azis','24060124140166',''),
    (8,'Muhammad Fauzan Akbar','24060124140139',''),
    (8,'Quinta Aurabiansyah','24060124120016',''),
    (9,'Akmal Dzaki Rahmatullah','24060124140151','Hill Climbing'),
    (9,'Wahyu Eko Setyo Pribowo','24060124120015',''),
    (10,'Muhammad Zaidaan Ardiyansyah','24060124140200',"Ruang's Games"),
    (10,'Muhammad Kemal Faza','24060124120013',''),
    (10,'Anintya Abhi Wiryateja','24060124130053',''),
    (11,'Mischa Nathanael Lumban Tobing','24060124140175','Caterpillar Game 3D'),
    (11,'Levi Ramot Siahaan','24060124130067',''),
]


# ============================================================
# 5. PARSE GTI D MEMBERS FROM VIDEO DESCRIPTION
# ============================================================
def parse_d_members(video_data):
    """Extract names & NIMs from GTI D descriptions, build student rows."""
    d_rows = {}
    for i in range(1, 11):
        key = f'D{i}'
        vd = video_data.get(key, {})
        desc = vd.get('desc', '')
        title = vd.get('title', '')
        vid = vd.get('vid', '')
        link = f'https://youtu.be/{vid}' if vid else '-'
        project = re.sub(r'(bts|behind the scene.*|kelompok.*|kelas.*)',
                         '', title, flags=re.IGNORECASE).strip(' -()_')

        # Extract NIM + Name pairs from description
        # Pattern: NIM (24xxxxxxxxxx) followed by name or vice versa
        members = []
        nim_name = re.findall(r'(24\d{10})\s*[-–:.]?\s*([A-Za-z][A-Za-z \'\.\(\)]*)', desc)
        for nim, name in nim_name:
            name = name.strip()
            if len(name) > 2 and not re.match(r'^(dan|dan|the|of|in|at|is)\b', name, re.I):
                members.append((name.title(), nim))

        # Also try reverse: Name then NIM
        if not members:
            name_nim = re.findall(r'([A-Z][a-zA-Z ]{3,30})\s*\(?(\d{14})\)?', desc)
            for name, nim in name_nim:
                members.append((name.strip().title(), nim))

        # Numbered list pattern: "1. Name NIM" or "1. Name (NIM)"
        if not members:
            lines = desc.split('\n')
            for line in lines:
                m = re.match(r'\d+[.)]\s*(.+?)\s*\(?(\d{14})\)?', line.strip())
                if m:
                    members.append((m.group(1).strip().title(), m.group(2)))

        d_rows[i] = {
            'members': members if members else [('-', '-')],
            'project': project,
            'link': link,
            'desc': desc,
            'title': title,
        }
    return d_rows

# ============================================================
# 6. VIDEO → KELOMPOK MAPPING
# ============================================================
KELOMPOK_VIDEO = {
    'A': {3:'A3',4:'A4',5:'A5',6:'A6',7:'A7',8:'A8',9:'A9',10:'A10',11:'A11',13:'A13'},
    'B': {1:'B1',2:'B2',3:'B3',4:'B4',5:'B5',6:'B6',7:'B7',8:'B8',9:'B9',10:'B10',11:'B11'},
    'C': {3:'C3',4:'C4',5:'C5',6:'C6',7:'C7',8:'C8',9:'C9',11:'C11',12:'C12'},
    'D': {1:'D1',2:'D2',3:'D3',4:'D4',5:'D5',6:'D6',7:'D7',8:'D8',9:'D9',10:'D10'},
    'E': {2:'E2',4:'E4',9:'E9'},
}


# ============================================================
# 7. EXCEL BUILDER
# ============================================================
def thin(): s=Side(style='thin'); return Border(left=s,right=s,top=s,bottom=s)
def med(): s=Side(style='medium'); return Border(left=s,right=s,top=s,bottom=s)
def fill(c): return PatternFill('solid',fgColor=c)
def fnt(bold=False,color='FF000000',sz=9): return Font(bold=bold,color=color,size=sz,name='Calibri')
def ca(wrap=True): return Alignment(horizontal='center',vertical='center',wrap_text=wrap)
def la(wrap=True): return Alignment(horizontal='left',vertical='center',wrap_text=wrap)

C_HDR='FF1F4E79'; C_HDR2='FF2E75B6'; C_WHITE='FFFFFFFF'; C_WARN='FFFFC7CE'
C_ODD='FFEEF4FB'; C_EVEN='FFFFFFFF'; C_TOT='FFFFF2CC'; C_GRD='FFE2EFDA'
C_SCORE_LOW='FFFFC7CE'; C_SCORE_MID='FFFFEB9C'; C_SCORE_HI='FFC6EFCE'

SCORE_COLS = [
    ('A','Kualitas\nVideo BTS\n(max 20)',20),
    ('B','Representasi\nObjek 3D\n(max 10)',10),
    ('C','Proyeksi\nOrtho\n(max 10)',10),
    ('D','Proyeksi\nPerspektif\n(max 10)',10),
    ('E','Geometri\n(max 8)',8),
    ('F','Kamera\n(max 8)',8),
    ('G','Cahaya\n(max 8)',8),
    ('H','Karakt.\nPermukaan\n(max 8)',8),
    ('I','Algoritma\nRendering\n(max 8)',8),
    ('J','Shading\n(max 10)',10),
]
HEADERS = ['No.\nKlp','NIM','Nama Mahasiswa','Nama Proyek','Link Video'] + \
          [s[1] for s in SCORE_COLS] + ['TOTAL\n(max 100)','Grade','Catatan Penilaian']

def sc(cell,bg=None,bold=False,color='FF000000',sz=9,align='c',border=True,wrap=True):
    if bg: cell.fill=fill(bg)
    cell.font=fnt(bold=bold,color=color,sz=sz)
    cell.alignment=ca(wrap) if align=='c' else la(wrap)
    if border: cell.border=thin()

def score_color(val, max_val):
    if val is None: return C_WHITE
    ratio = val / max_val
    if ratio >= 0.75: return C_SCORE_HI
    elif ratio >= 0.55: return C_SCORE_MID
    else: return C_SCORE_LOW

def build_sheet(ws, kelas, student_rows, video_data, kelompok_video_map):
    """
    student_rows: list of (klp, nama, nim, proyek)
    kelompok_video_map: {klp_no: video_key}
    """
    ws.freeze_panes = 'A4'
    # Title row
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(HEADERS))
    t=ws.cell(1,1); t.value=f'PENILAIAN TUGAS BESAR GTI {kelas} — SEMESTER GENAP 2025/2026'
    sc(t,bg=C_HDR,bold=True,color=C_WHITE,sz=13)
    ws.row_dimensions[1].height=22

    # Header row
    for ci,h in enumerate(HEADERS,1):
        cell=ws.cell(2,ci); cell.value=h
        if ci<=5: bg=C_HDR
        elif ci==len(HEADERS)-2: bg='FFBF8F00'
        elif ci>=len(HEADERS)-1: bg='FF375623'
        else: bg=C_HDR2
        sc(cell,bg=bg,bold=True,color=C_WHITE,sz=8)
    ws.row_dimensions[2].height=52

    # Sub-header: max bobot
    sub = ['']*5 + [str(s[2]) for s in SCORE_COLS] + ['100','','']
    for ci,v in enumerate(sub,1):
        cell=ws.cell(3,ci); cell.value=v if v else ''
        sc(cell,bg='FFD9E1F2',bold=True,color='FF1F3864',sz=8)
    ws.row_dimensions[3].height=14

    # Column widths
    widths=[8,18,30,24,42]+[12]*10+[12,7,45]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    prev_klp=None; toggle=0; row_num=3
    for (klp,nama,nim,proyek) in student_rows:
        row_num += 1
        if klp!=prev_klp: toggle=1-toggle; prev_klp=klp
        row_bg=C_ODD if toggle else C_EVEN
        no_data = (nama=='-' or nama=='')

        vid_key = kelompok_video_map.get(klp)
        vd = video_data.get(vid_key,{}) if vid_key else {}
        title = vd.get('title','')
        desc = vd.get('desc','')
        vid_id = vd.get('vid','')
        link = f'https://youtu.be/{vid_id}' if vid_id else '-'

        # Score this kelompok
        s = score_video(vid_key, title, desc) if vid_key and (title or desc) else None
        no_video = (not vid_key) or (not s)

        # Get project name from title if missing
        proj = proyek
        if not proj or proj=='-':
            # Try extract from title
            p = re.sub(r'\s*(bts|behind the scene.*|proyek.*|penugasan.*|kelompok.*)',
                       '', title, flags=re.IGNORECASE).strip(' -()')
            proj = p if p else title

        # Write identity cols
        for ci,v in enumerate([klp,nim,nama,proj,link],1):
            cell=ws.cell(row_num,ci)
            cell.value=str(v)
            cell.fill=fill(C_WARN if no_data else row_bg)
            if ci==5 and str(v).startswith('http'):
                cell.hyperlink=str(v)
                cell.font=Font(name='Calibri',size=9,color='FF0563C1',underline='single')
            else:
                cell.font=fnt(sz=9,color='FF000000')
            cell.alignment=ca() if ci==1 else la()
            cell.border=thin()

        # Write score cols (6-15)
        score_keys=['A','B','C','D','E','F','G','H','I','J']
        note_parts=[]
        for ci, (sk, _, max_v) in enumerate(SCORE_COLS, start=6):
            cell=ws.cell(row_num,ci)
            if no_data or no_video:
                cell.value=None
                cell.fill=fill(C_WARN if no_data else 'FFFFD7BE')
            else:
                val=s[sk]
                cell.value=val
                cell.fill=fill(score_color(val, max_v))
                if val < max_v * 0.5:
                    note_parts.append(sk)
            cell.font=fnt(bold=True if s and not no_data else False, sz=9)
            cell.alignment=ca()
            cell.border=thin()

        # Total col (16)
        tot_cell=ws.cell(row_num,16)
        if no_data or no_video:
            tot_cell.value=None
            tot_cell.fill=fill(C_WARN if no_data else 'FFFFD7BE')
        else:
            tot_cell.value=s['total']
            tot_cell.fill=fill(score_color(s['total'],100))
            tot_cell.font=fnt(bold=True,sz=10)
        tot_cell.alignment=ca(); tot_cell.border=thin()

        # Grade col (17)
        grd_cell=ws.cell(row_num,17)
        if s and not no_data:
            grd_cell.value=s['grade']
            grd_cell.fill=fill(C_SCORE_HI if s['total']>=75 else C_SCORE_MID if s['total']>=55 else C_SCORE_LOW)
            grd_cell.font=fnt(bold=True,sz=10)
        else:
            grd_cell.fill=fill(C_WARN if no_data else 'FFFFD7BE')
        grd_cell.alignment=ca(); grd_cell.border=thin()

        # Catatan col (18)
        cat_cell=ws.cell(row_num,18)
        if no_data:
            cat_cell.value='Belum mengumpulkan / data tidak tersedia'
            cat_cell.fill=fill(C_WARN)
        elif no_video:
            cat_cell.value='Video tidak dapat diakses (404/private)'
            cat_cell.fill=fill('FFFFD7BE')
        else:
            weak = ', '.join(note_parts)
            cat_str = f'Sumber: deskripsi YouTube. Total={s["total"]}.'
            if weak: cat_str += f' Komponen lemah: {weak}.'
            cat_cell.value=cat_str
            cat_cell.fill=fill(row_bg)
        cat_cell.font=fnt(sz=8); cat_cell.alignment=la(wrap=True); cat_cell.border=thin()
        ws.row_dimensions[row_num].height=18

    # Disclaimer row
    disc_row = row_num+2
    ws.merge_cells(start_row=disc_row,start_column=1,end_row=disc_row,end_column=len(HEADERS))
    d=ws.cell(disc_row,1)
    d.value=('⚠️  Nilai dihitung otomatis berdasarkan analisis deskripsi video YouTube. '
             'Disarankan dosen/asisten mereview langsung video sebelum finalisasi. '
             'Skala: A≥85 | AB 75-84 | B 65-74 | BC 55-64 | C 45-54 | D 35-44 | E<35')
    d.fill=fill('FFFCE4D6'); d.font=fnt(bold=True,color='FF833C00',sz=9)
    d.alignment=la(); d.border=thin(); ws.row_dimensions[disc_row].height=22


# ============================================================
# 8. BUILD RUBRIK SHEET
# ============================================================
def build_rubrik_sheet(ws):
    ws.column_dimensions['A'].width=8; ws.column_dimensions['B'].width=26
    ws.column_dimensions['C'].width=55; ws.column_dimensions['D'].width=14
    ws.column_dimensions['E'].width=55
    ws.merge_cells('A1:E1')
    t=ws.cell(1,1); t.value='RUBRIK PENILAIAN — GRAFIKA & TEKNOLOGI INFORMASI (GTI) 2026'
    sc(t,bg=C_HDR,bold=True,color=C_WHITE,sz=13)
    ws.merge_cells('A2:E2')
    s=ws.cell(2,1); s.value='Tugas: Video Behind The Scenes (BTS) ≤ 5 Menit | Upload YouTube'
    sc(s,bg=C_HDR2,bold=True,color=C_WHITE,sz=11)
    for ci,h in enumerate(['Kode','Komponen','Deskripsi & Indikator','Bobot','Kriteria Skor'],1):
        c=ws.cell(3,ci); c.value=h; sc(c,bg=C_HDR2,bold=True,color=C_WHITE,sz=9)
    ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=18; ws.row_dimensions[3].height=18

    rubrik_rows=[
        ('A','Kualitas Video BTS',
         'Durasi ≤5 mnt; judul format [Proyek] BTS ([Panggilan]); deskripsi berisi nama+NIM semua anggota + penjelasan singkat proyek; video dapat diakses',20,
         '18-20: Semua komponen terpenuhi\n14-17: Sebagian besar terpenuhi\n10-13: Beberapa kurang\n<10: Banyak tidak terpenuhi'),
        ('B','Representasi Objek 3D',
         'Objek 3D divisualisasikan menggunakan teknik yang benar: primitif (kubus, bola, kerucut, silinder), mesh polygon, vertex, edge, face',10,
         '9-10: Representasi lengkap & jelas\n7-8: Ditunjukkan, sedikit kurang\n5-6: Ada tapi tidak dijelaskan\n3-4: Sangat minim'),
        ('C','Proyeksi Orthographic',
         'Proyeksi ortographic (tampak atas/depan/samping, parallel projection) diterapkan dan dijelaskan secara eksplisit',10,
         '9-10: Diterapkan & dijelaskan sangat jelas\n7-8: Diterapkan, penjelasan cukup\n5-6: Ada tapi kurang\n3-4: Tidak ada/sangat minim'),
        ('D','Proyeksi Perspektif',
         'Perspektif 1-point, 2-point, atau 3-point diterapkan. Semakin banyak jenis perspektif, nilai lebih tinggi',10,
         '9-10: Multi-perspektif & sangat jelas\n7-8: Satu perspektif, jelas\n5-6: Disebutkan, kurang detail\n3-4: Tidak ada/sangat minim'),
        ('E','Geometri',
         'Transformasi geometri: translasi, rotasi, skala/scale, shear — diterapkan pada objek 3D dalam proyek',8,
         '7-8: Transformasi lengkap & jelas\n5-6: Transformasi dasar ada\n3-4: Sebagian ada\n1-2: Tidak ada'),
        ('F','Kamera',
         'Setup kamera yang benar: FOV, near/far clipping plane, viewport, posisi/orientasi (lookat)',8,
         '7-8: Setup kamera lengkap & dijelaskan\n5-6: Sebagian terpenuhi\n3-4: Ada kamera, tidak dijelaskan\n1-2: Tidak ada'),
        ('G','Cahaya',
         'Pencahayaan: minimal 2 dari: ambient, diffuse, specular, directional light, point light, spot light',8,
         '7-8: ≥2 jenis cahaya, terlihat & dijelaskan\n5-6: 1 jenis cahaya\n3-4: Ada cahaya, tidak dijelaskan\n1-2: Tidak ada'),
        ('H','Karakteristik Permukaan',
         'Material/tekstur: warna, roughness, metallic, emissive, atau tekstur gambar — diterapkan pada objek',8,
         '7-8: Material kaya & beragam\n5-6: Material dasar ada\n3-4: Warna polos saja\n1-2: Tidak ada'),
        ('I','Algoritma Rendering',
         'Algoritma rendering dijelaskan: wireframe, scanline rasterization, depth buffer, z-buffer, atau pipeline OpenGL/GLUT',8,
         '7-8: Algoritma dijelaskan & diterapkan\n5-6: Disebutkan, kurang detail\n3-4: Tersirat, tidak eksplisit\n1-2: Tidak ada'),
        ('J','Shading',
         'Model shading diterapkan & dijelaskan: flat shading, Gouraud shading, Phong shading, atau variasi lainnya',10,
         '9-10: Shading tepat, jelas, detail\n7-8: Shading ada, cukup\n5-6: Shading ada, tidak dijelaskan\n3-4: Tidak ada/sangat minim'),
    ]
    for ri,(kode,komp,indikator,bobot,kriteria) in enumerate(rubrik_rows,4):
        bg=C_ODD if ri%2==0 else C_EVEN
        for ci,v in enumerate([kode,komp,indikator,bobot,kriteria],1):
            c=ws.cell(ri,ci); c.value=v
            c.fill=fill(C_TOT if ci==4 else bg)
            c.font=fnt(bold=(ci in(1,4)),color='FF833C00' if ci==4 else 'FF000000',sz=8)
            c.alignment=ca() if ci in(1,4) else la()
            c.border=thin()
        ws.row_dimensions[ri].height=65
    tot_r=len(rubrik_rows)+4
    ws.merge_cells(start_row=tot_r,start_column=1,end_row=tot_r,end_column=3)
    tot=ws.cell(tot_r,1); tot.value='TOTAL BOBOT KESELURUHAN'
    sc(tot,bg=C_HDR,bold=True,color=C_WHITE,sz=11)
    tv=ws.cell(tot_r,4); tv.value=100
    sc(tv,bg='FFBF8F00',bold=True,color=C_WHITE,sz=12)
    ws.row_dimensions[tot_r].height=22


# ============================================================
# 9. MAIN
# ============================================================
def main():
    OUT = '/projects/sandbox/sonet-product/nilai-nilai/gti-2026/Penilaian_GTI_2026_SCORED.xlsx'

    # Fetch all YouTube data
    video_data = fetch_all()

    # Save raw data
    with open('/projects/sandbox/sonet-product/nilai-nilai/gti-2026/video_descriptions.json','w',encoding='utf-8') as f:
        json.dump(video_data, f, ensure_ascii=False, indent=2)
    print(f"\nSaved {len(video_data)} video records.")

    # Parse GTI D members from descriptions
    d_rows_parsed = parse_d_members(video_data)
    DATA_GTI_D = []
    for i in range(1,11):
        info = d_rows_parsed[i]
        members = info['members']
        project = info['project']
        for idx,(nm,nim) in enumerate(members):
            DATA_GTI_D.append((i, nm, nim, project))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_r = wb.create_sheet('RUBRIK', 0)
    build_rubrik_sheet(ws_r)
    print("✓ Sheet RUBRIK")

    sheets = [
        ('GTI A', DATA_GTI_A, 'A'),
        ('GTI B', DATA_GTI_B, 'B'),
        ('GTI C', DATA_GTI_C, 'C'),
        ('GTI D', DATA_GTI_D, 'D'),
        ('GTI E', DATA_GTI_E, 'E'),
    ]

    for sheet_name, sdata, kelas in sheets:
        ws = wb.create_sheet(sheet_name)
        kvm = KELOMPOK_VIDEO[kelas]
        build_sheet(ws, sheet_name.split()[-1], sdata, video_data, kvm)
        # Count scored
        scored = sum(1 for (klp,nm,nim,p) in sdata if nm!='-' and nm!='' and klp in kvm)
        print(f"✓ Sheet {sheet_name}: {len(sdata)} baris, {scored} kelompok berhasil dinilai")

    wb.save(OUT)
    print(f"\n✅ File tersimpan: {OUT}")

    # Print scoring summary
    print("\n=== RINGKASAN NILAI ===")
    for sheet_name, sdata, kelas in sheets:
        kvm = KELOMPOK_VIDEO[kelas]
        seen = set()
        for (klp,nm,nim,p) in sdata:
            if nm=='-' or nm=='' or klp in seen: continue
            seen.add(klp)
            vkey = kvm.get(klp)
            if vkey:
                vd = video_data.get(vkey,{})
                s = score_video(vkey, vd.get('title',''), vd.get('desc',''))
                if s:
                    print(f"  {kelas}{klp}: {vd.get('title','')[:40]} → Total={s['total']} ({s['grade']})")

if __name__ == '__main__':
    main()
