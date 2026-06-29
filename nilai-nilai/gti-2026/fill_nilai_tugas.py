#!/usr/bin/env python3
"""
fill_nilai_tugas.py
───────────────────
1. Baca scored Excel → bangun NIM → skor kelompok
2. Cari mahasiswa yg belum ada skornya → retry YouTube fetch
3. Pastikan semua skor ≥ 60
4. Isi kolom Nilai Tugas (G) di setiap file GTI PAIK/MIK
5. Update Penilaian_GTI_2026_SCORED.xlsx dengan nilai final
"""

import openpyxl, glob, json, re, os, requests
from openpyxl.styles import PatternFill, Font, Alignment
from concurrent.futures import ThreadPoolExecutor, as_completed

BASE = '/projects/sandbox/sonet-product/nilai-nilai/gti-2026'
SCORED_FILE   = f'{BASE}/Penilaian_GTI_2026_SCORED.xlsx'
FINAL_FILE    = f'{BASE}/Penilaian_GTI_2026_FINAL.xlsx'
DESC_FILE     = f'{BASE}/video_descriptions.json'

MIN_SCORE = 60  # nilai minimum

# ─────────────────────────────────────────────
# STEP 1 – Baca scored Excel, build NIM→score
# ─────────────────────────────────────────────
def build_nim_score_map(scored_path):
    wb = openpyxl.load_workbook(scored_path)
    # Map: NIM → {'total', 'grade', 'sheet', 'klp', 'nama'}
    # Also map: (sheet_kelas, klp) → {'total','grade'}
    nim_map   = {}   # NIM string → score dict
    group_map = {}   # (kelas_letter, klp_no) → score dict

    for sname in wb.sheetnames:
        if sname == 'RUBRIK': continue
        ws = wb[sname]
        kelas = sname.split()[-1]  # 'A', 'B', ...
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[1]: continue
            klp  = row[0]
            nim  = str(row[1]).strip()
            nama = str(row[2]).strip() if row[2] else ''
            total = row[15]
            grade = row[16]
            if nim in ('-', '', 'NIM') or len(nim) < 10: continue

            raw_total = total if isinstance(total, (int, float)) else 0
            final_total = max(MIN_SCORE, round(raw_total)) if raw_total > 0 else 0
            # 0 means no video at all; we'll handle later

            nim_map[nim] = {
                'total': final_total,
                'raw_total': round(raw_total) if raw_total else 0,
                'grade': grade,
                'kelas': kelas,
                'klp': klp,
                'nama': nama,
                'has_score': raw_total > 0,
            }
            gkey = (kelas, klp)
            if gkey not in group_map or raw_total > 0:
                group_map[gkey] = {'total': final_total, 'raw': round(raw_total)}

    print(f"NIM mapped from scored Excel: {len(nim_map)}")
    no_score = [nim for nim, v in nim_map.items() if not v['has_score']]
    print(f"  → No score (video missing/empty): {len(no_score)}")
    return nim_map, group_map, wb


# ─────────────────────────────────────────────
# STEP 2 – Cari semua NIM di GTI files yg belum di nim_map
# ─────────────────────────────────────────────
def get_all_gti_students():
    """Return dict NIM → {'nama', 'file', 'kelas', 'prodi', 'row'}"""
    students = {}
    files = sorted(glob.glob(f'{BASE}/GTI_*.xlsx'))
    for fpath in files:
        wb = openpyxl.load_workbook(fpath)
        ws = wb.active
        fname = os.path.basename(fpath)
        prodi = 'MIK' if 'MIK' in fname else 'PAIK'
        kelas = fname.split('_')[-1].replace('.xlsx','')
        for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
            nim = str(row[0]).strip() if row[0] else ''
            nama = str(row[1]).strip() if row[1] else ''
            if not nim or nim == 'NIM' or len(nim) < 10: continue
            if nim not in students:
                students[nim] = {'nama': nama, 'file': fname, 'kelas': kelas, 'prodi': prodi}
    print(f"\nTotal students in GTI PAIK/MIK files: {len(students)}")
    return students


# ─────────────────────────────────────────────
# STEP 3 – Retry YouTube untuk NIM yang belum ada nilainya
# ─────────────────────────────────────────────
RUBRIK_KEYWORDS = {
    'B': ['vertex','edge','face','mesh','polygon','primitif','primitive','kubus','bola','kerucut','silinder','cylinder','cube','sphere','cone','bangun 3d','objek 3d','3d object','geometri 3d'],
    'C': ['orthographic','ortho','ortografik','tampak atas','tampak depan','tampak samping','top view','front view','side view','parallel projection','proyeksi sejajar'],
    'D': ['perspektif','perspective','1 point','2 point','3 point','one point','two point','vanishing','titik hilang','frustum','fov','field of view','proyeksi perspektif'],
    'E': ['translasi','rotasi','skala','transformasi','translation','rotation','scale','transform','matrix','glrotatef','gltranslatef','glscalef','translate','rotate'],
    'F': ['kamera','camera','fov','field of view','viewport','clipping','frustum','lookat','look at','near plane','far plane','gluperspective','glulookat','view matrix'],
    'G': ['cahaya','lighting','ambient','diffuse','specular','shininess','directional light','point light','spot light','pencahayaan','iluminasi','gllightfv','gl_light','gl_ambient'],
    'H': ['material','texture','tekstur','roughness','metallic','permukaan','surface','glmaterialfv','reflection','refleksi','specular map','normal map','bump map','emissive','albedo','uv mapping'],
    'I': ['rendering','rasterization','rasterisasi','pipeline','wireframe','scanline','opengl','glut','fragment shader','vertex shader','depth buffer','z-buffer','hidden surface'],
    'J': ['shading','flat shading','gouraud','phong','normal map','shader','smooth shading','constant shading','glshademodel','gl_smooth','gl_flat'],
}
MAX_SCORES = {'A':20,'B':10,'C':10,'D':10,'E':8,'F':8,'G':8,'H':8,'I':8,'J':10}

def score_A(title, desc):
    score = 0
    lower = (title + ' ' + desc).lower()
    score += 5 if re.search(r'bts', title.lower()) else (4 if 'behind the scene' in title.lower() else 2)
    nim_count = len(re.findall(r'240\d{9}', desc))
    score += 5 if nim_count>=3 else (4 if nim_count>=2 else (3 if nim_count>=1 else 1))
    tech_words = ['teknik','rendering','proyeksi','projection','kamera','camera','cahaya','lighting','material','texture','opengl','3d','transformasi','shading']
    tech_hits = sum(1 for w in tech_words if w in lower)
    score += 5 if tech_hits>=5 else (4 if tech_hits>=3 else (3 if tech_hits>=1 else 1))
    score += 5 if desc else 0
    return min(score, 20)

def score_comp(text, key, max_v):
    kws = RUBRIK_KEYWORDS[key]
    hits = sum(1 for k in kws if k in text)
    ratio = hits / len(kws)
    if ratio >= 0.4: base = 0.90
    elif ratio >= 0.2: base = 0.78
    elif ratio >= 0.1: base = 0.65
    elif hits > 0:    base = 0.55
    else:             base = 0.42
    return round(max_v * base)

def score_video_from_text(title, desc):
    text = (title + '\n' + desc).lower()
    scores = {'A': score_A(title, desc)}
    for k, mv in MAX_SCORES.items():
        if k != 'A': scores[k] = score_comp(text, k, mv)
    total = sum(scores.values())
    return max(MIN_SCORE, total)

def fetch_yt_desc(vid_id, session):
    try:
        r = session.get(f'https://www.youtube.com/watch?v={vid_id}', timeout=15)
        text = r.text
        desc = ''
        m = re.search(r'"attributedDescriptionBodyText":\{"content":"(.*?)"(?:,"styleRuns"|,"commandRuns"|\})', text, re.DOTALL)
        if m: desc = m.group(1).replace('\\n','\n').replace('\\"','"')
        if not desc:
            m2 = re.search(r'"shortDescription":"((?:[^"\\]|\\.)*?)","isCrawlable"', text)
            if m2: desc = m2.group(1).replace('\\n','\n').replace('\\"','"')
        try:
            ro = requests.get(f'https://www.youtube.com/oembed?url=https://www.youtube.com/watch?v={vid_id}&format=json', timeout=6)
            title = ro.json().get('title','')
        except: title = ''
        return title, desc
    except: return '', ''


# ─────────────────────────────────────────────
# STEP 4 – Isi Nilai Tugas di semua GTI files
# ─────────────────────────────────────────────
def fill_nilai_tugas(nim_final_map):
    """Fill column G (Nilai Tugas) in all GTI PAIK/MIK files."""
    files = sorted(glob.glob(f'{BASE}/GTI_*.xlsx'))
    summary = []

    for fpath in files:
        fname = os.path.basename(fpath)
        wb = openpyxl.load_workbook(fpath)
        ws = wb.active
        filled = 0; not_found = 0

        for i, row in enumerate(ws.iter_rows(min_row=7), start=7):
            nim_cell = row[0]
            nim = str(nim_cell.value).strip() if nim_cell.value else ''
            if not nim or nim == 'NIM' or len(nim) < 10: continue

            score_data = nim_final_map.get(nim)
            if score_data:
                val = score_data['final']
                filled += 1
            else:
                val = MIN_SCORE  # default min
                not_found += 1

            tugas_cell = ws.cell(row=i, column=7)  # Column G
            tugas_cell.value = val
            # Highlight sesuai nilai
            if val >= 80:    bg = 'FFC6EFCE'
            elif val >= 70:  bg = 'FFEBF1DE'
            elif val >= 60:  bg = 'FFFFEB9C'
            else:            bg = 'FFFFC7CE'
            tugas_cell.fill = PatternFill('solid', fgColor=bg)
            tugas_cell.font = Font(name='Calibri', size=10, bold=True)
            tugas_cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(fpath)
        summary.append({'file': fname, 'filled': filled, 'default': not_found})
        print(f"  ✓ {fname}: {filled} terisi dari scored, {not_found} default={MIN_SCORE}")
    return summary


# ─────────────────────────────────────────────
# STEP 5 – Update Scored Excel dengan nilai final
# ─────────────────────────────────────────────
def update_scored_excel(wb_scored, nim_final_map):
    """Update total & grade di scored Excel dengan nilai yang sudah fixed."""
    for sname in wb_scored.sheetnames:
        if sname == 'RUBRIK': continue
        ws = wb_scored[sname]
        for row in ws.iter_rows(min_row=4):
            nim_cell = row[1]
            nim = str(nim_cell.value).strip() if nim_cell.value else ''
            if not nim or nim == '-' or nim == 'NIM' or len(nim) < 10: continue
            data = nim_final_map.get(nim)
            if data:
                # Update total (col 16 = index 15)
                tot_cell = row[15]
                grd_cell = row[16]
                cat_cell = row[17]
                tot_cell.value = data['final']
                grd_cell.value = data['grade_final']
                if data.get('raised'):
                    if cat_cell.value:
                        cat_cell.value = str(cat_cell.value) + f' [dinaikkan dari {data["raw"]} → {data["final"]}]'
                    else:
                        cat_cell.value = f'Nilai dinaikkan ke minimum {MIN_SCORE}'
                # Color update
                score = data['final']
                if score >= 85: bg = 'FFC6EFCE'
                elif score >= 75: bg = 'FFEBF1DE'
                elif score >= 65: bg = 'FFFFEB9C'
                else: bg = 'FFFFD7BE'
                for c in [tot_cell, grd_cell]:
                    c.fill = PatternFill('solid', fgColor=bg)
                    c.font = Font(name='Calibri', bold=True, size=10)
    wb_scored.save(FINAL_FILE)
    print(f"\n✅ Final scored Excel saved: {FINAL_FILE}")


def grade_from_total(t):
    if t >= 85: return 'A'
    elif t >= 75: return 'AB'
    elif t >= 65: return 'B'
    elif t >= 55: return 'BC'
    elif t >= 45: return 'C'
    elif t >= 35: return 'D'
    else: return 'E'


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("GTI 2026 — Fill Nilai Tugas & Fix Min Score")
    print("=" * 60)

    # Load existing video descriptions if available
    existing_desc = {}
    if os.path.exists(DESC_FILE):
        with open(DESC_FILE) as f:
            existing_desc = json.load(f)
        print(f"Loaded {len(existing_desc)} existing video descriptions")

    # Step 1: Build NIM → score from scored Excel
    nim_map, group_map, wb_scored = build_nim_score_map(SCORED_FILE)

    # Step 2: Get all students from GTI files
    all_students = get_all_gti_students()

    # Step 3: Find students NOT in scored Excel and not in any group
    missing_nims = [nim for nim in all_students if nim not in nim_map]
    print(f"\nStudents in GTI files but NOT in scored Excel: {len(missing_nims)}")
    for nim in missing_nims:
        print(f"  {nim}: {all_students[nim]['nama']}")

    # Step 4: For missing students, try to find their group from video descriptions
    # by matching NIM in any description
    nim_to_video = {}
    for vkey, vdata in existing_desc.items():
        desc = vdata.get('desc', '')
        nims_in_desc = re.findall(r'(24\d{10})', desc)
        for nim in nims_in_desc:
            if nim not in nim_to_video:
                nim_to_video[nim] = vkey

    # Score missing NIMs using their video's description
    extra_scores = {}
    for nim in missing_nims:
        vkey = nim_to_video.get(nim)
        if vkey and vkey in existing_desc:
            vd = existing_desc[vkey]
            total = score_video_from_text(vd.get('title',''), vd.get('desc',''))
            extra_scores[nim] = {'total': total, 'source': f'video_{vkey}'}
            print(f"  Found via desc: {nim} → video {vkey} → score {total}")
        else:
            # For PAIK students who are older cohort (2021, 2022, 2023 NIM)
            # They might be in groups not captured, give reasonable default
            extra_scores[nim] = {'total': 0, 'source': 'not_found'}

    # Try YouTube fetch for still-missing NIMs
    still_missing_vids = {}
    # Check PAIK files for older students who might have different videos
    # GTI_PAIK6404 has students like 24060123*, 24060122*, 24060121*
    # These are older cohort - check if they're in any group video
    paik_missing = [nim for nim in missing_nims if extra_scores.get(nim,{}).get('total',0) == 0]
    print(f"\nStill missing after desc search: {len(paik_missing)}")

    # For PAIK students, try fetching from the group they belong to
    # by checking PAIK class groupings
    # We'll try re-fetching PAIK B which has NIM from 2022 cohort
    # These students (Ananda Bagus 24060122130091, etc.) are in GTI B kelompok 7
    # Map from known data
    KNOWN_PAIK_GROUPS = {
        '24060121140176': 'A5',   # Fernanda Galih - GTI A klp 5
        '24060122130091': 'B7',   # Ananda Bagus - GTI B klp 7
        '24060122140141': 'B7',   # Hafidz Zulfikar - GTI B klp 7
        '24060122140147': 'B7',   # Nabil Razaki - GTI B klp 7
        '24060122140122': 'B7',   # Rafi Deandra - GTI B klp 7
        '24060110142': 'E7',      # Muchamad Yuda - GTI E klp 7 (partial NIM)
        '24060124110142': 'E7',   # Muchamad Yuda - GTI E klp 7
    }

    # Step 5: Build FINAL NIM map (scored + extra + raised min 60)
    nim_final_map = {}

    # From scored Excel
    for nim, data in nim_map.items():
        raw = data['raw_total']
        final = max(MIN_SCORE, raw) if raw > 0 else 0
        raised = (raw > 0 and raw < MIN_SCORE)
        nim_final_map[nim] = {
            'final': final if final > 0 else MIN_SCORE,
            'raw': raw,
            'grade_final': grade_from_total(max(MIN_SCORE, raw) if raw > 0 else MIN_SCORE),
            'raised': raised,
            'source': 'scored_excel',
        }

    # From extra searches
    for nim in missing_nims:
        # Check KNOWN_PAIK_GROUPS first
        vkey = KNOWN_PAIK_GROUPS.get(nim) or nim_to_video.get(nim)
        if vkey and vkey in existing_desc:
            vd = existing_desc[vkey]
            total = score_video_from_text(vd.get('title',''), vd.get('desc',''))
            final = max(MIN_SCORE, total)
        else:
            final = MIN_SCORE  # Default min for unfound

        nim_final_map[nim] = {
            'final': final,
            'raw': final,
            'grade_final': grade_from_total(final),
            'raised': False,
            'source': f'found_via_{vkey}' if vkey else 'default_min',
        }

    print(f"\nFinal NIM map: {len(nim_final_map)} entries")
    raised_count = sum(1 for v in nim_final_map.values() if v.get('raised'))
    default_count = sum(1 for v in nim_final_map.values() if v['source'].startswith('default'))
    print(f"  → Raised to ≥60: {raised_count}")
    print(f"  → Default min 60: {default_count}")

    # Step 6: Fill Nilai Tugas in all GTI files
    print("\n=== Filling Nilai Tugas ===")
    summary = fill_nilai_tugas(nim_final_map)

    # Step 7: Update scored Excel with final values
    print("\n=== Updating scored Excel ===")
    update_scored_excel(wb_scored, nim_final_map)

    # Print full summary
    print("\n" + "=" * 60)
    print("RINGKASAN NILAI TUGAS AKHIR")
    print("=" * 60)
    print(f"\n{'File':<45} {'Terisi':>6} {'Default':>8}")
    print("-" * 62)
    for s in summary:
        print(f"  {s['file']:<43} {s['filled']:>6} {s['default']:>8}")

    print(f"\n{'NIM':<18} {'Nama':<35} {'Nilai':>6} {'Sumber'}")
    print("-" * 75)
    # Print all with score < 70 for review
    for nim, v in sorted(nim_final_map.items()):
        nama = all_students.get(nim, nim_map.get(nim, {}).get('nama', '?') if nim in nim_map else '?')
        if isinstance(nama, dict): nama = nama.get('nama', '?')
        if v['final'] < 70:
            print(f"  {nim:<16} {nama[:33]:<35} {v['final']:>6}  {v['source']}")

    print("\n✅ Selesai!")

if __name__ == '__main__':
    main()
