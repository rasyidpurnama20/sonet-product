import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

os.makedirs('/projects/sandbox/sonet-product/nilai-nilai/gti-2026', exist_ok=True)

# ============================================================
# DATA MAHASISWA PER KELAS
# (kelompok, nama, nim, nama_proyek, link_video)
# ============================================================

DATA_GTI_A = [
    (1,  '-', '-', '-', '-'),
    (2,  '-', '-', '-', '-'),
    (3,  'Johan Reinhart Calvin',         '24060124130076', '-', 'https://youtu.be/Im8KWkQt2oA'),
    (3,  'Ilham Muhammad Raffi',           '24060124140198', '-', ''),
    (3,  'Romualdus Yoas Wicaksono',       '24060124120046', '-', ''),
    (3,  'Muhammad Akmal Fazli Riyadi',    '24060124130123', '-', ''),
    (4,  'Muhammad Fikri',                 '24060124130069', '-', 'https://youtu.be/_jsEwoba0yQ'),
    (4,  'Khanza Qaila',                   '24060124120011', '-', ''),
    (4,  'Nayla Husna',                    '24060124140158', '-', ''),
    (5,  'Shalom Kurniawan',               '24060124120033', '-', 'https://youtu.be/snYwaR38XQ4'),
    (5,  'Arsy Thariq Munawar',            '24060124120009', '-', ''),
    (5,  'Yuma Hazza Yuditama',            '24060124120035', '-', ''),
    (5,  'Fernanda Galih Saputra',         '24060121140176', '-', ''),
    (6,  'Silvani Salsabilla',             '24060124130066', '-', 'https://youtu.be/Z59d6nk4PAs'),
    (6,  'Eileen Albert Tandrio',          '24060124140180', '-', ''),
    (6,  'Jessica Laurencia Panjaitan',    '24060124140084', '-', ''),
    (7,  'Muhammad Fahri',                 '24060124120037', '-', 'https://youtu.be/oQcXjDnyGJU'),
    (7,  'Nouvella Rahma Fitrah Legarsi',  '24060124120029', '-', ''),
    (7,  'Ovilia Suci Ramadhani',          '24060124120040', '-', ''),
    (7,  'Biyani Andarisky Maratia',       '24060124130070', '-', ''),
    (8,  'Agung Rama Pramana Putra',       '24060124120007', '-', 'https://youtu.be/f-mvyIQzta4'),
    (8,  'Andiny Khaerany Suhartady',      '24060124140194', '-', ''),
    (8,  'Hadrian Shandhy Yudha',          '24060124140207', '-', ''),
    (8,  'Mohammad Najib Fitrianto',       '24060124140203', '-', ''),
    (9,  'Elang Fadila Ahmad',             '24060124130108', '-', 'https://youtu.be/donL48coYQE'),
    (9,  'Dehar Zaidan Dzaki Amirullah',   '24060124130099', '-', ''),
    (9,  'Ridho Tri Saputra',              '24060124130122', '-', ''),
    (9,  'Axel Anggian Hamonangan Purba',  '24060124140127', '-', ''),
    (10, 'Ammar Rozan Rusyaidan',          '24060124140147', '-', 'https://youtu.be/NOa6aFGw6cs'),
    (10, 'Arya Naufal Akmal',              '24060124130089', '-', ''),
    (10, 'Misbachul Munir',                '24060124120031', '-', ''),
    (11, "Dzaki Fathul'Alim Cahyo",        '24060124130103', '-', 'https://youtu.be/PEcKOSPsxIc'),
    (11, 'Daniel Lamganda Tua Gultom',     '24060124120048', '-', ''),
    (11, 'Putri Elizabeth Simanjuntak',    '24060124120018', '-', ''),
    (12, '-', '-', '-', '-'),
    (13, 'Revanska Athallah Muhammad',     '24060124140129', '-', 'https://youtu.be/FltKiLTQOOE'),
    (13, 'Kiyoshi Akila Tira',             '24060124130074', '-', ''),
    (13, 'Yustinus Hendi Setyawan',        '24060124130114', '-', ''),
]


DATA_GTI_B = [
    (1,  'Muchammad Rajib Tafrichan',              '24060124140141', 'Car Drifting Animation',    'https://youtu.be/8CHXPfwOpLE'),
    (1,  'Fazl Nizam Priyambodho',                 '24060124130121', '', ''),
    (1,  'Muhammad Ibrahim Alghifari',             '24060124140140', '', ''),
    (1,  'Adam Mulya Rasyid',                      '24060124140179', '', ''),
    (2,  'Moses Morell Yosefan',                   '24060124130094', 'Furiosa Riding',             'https://youtu.be/ke7Sop09qNc'),
    (2,  'Lintang Aulia Nuraini',                  '24060124120017', '', ''),
    (2,  'Alyssa Shane Kurniawan',                 '24060124120038', '', ''),
    (2,  'Olivia Oktaviani',                       '24060124120050', '', ''),
    (3,  'Shafa Aqilla Zahira',                    '24060124140146', 'Look Around',                'https://youtu.be/kPP7qwyveFU'),
    (3,  'Rafi Anandra Dharmawan',                 '24060124130071', '', ''),
    (3,  'Shofwan Fikrul Huda',                    '24060124130106', '', ''),
    (4,  'Birela Miadeta Purita',                  '24060124120002', '3D TilesRush',               'https://youtu.be/w0AvKSJOdfg'),
    (4,  'Diah Maulida Pratiwi',                   '24060124120034', '', ''),
    (4,  "Hana Nafi'atul Haq",                     '24060124130081', '', ''),
    (4,  'Arga Yura Danendra',                     '24060124140191', '', ''),
    (5,  'Jordan Tenggara',                        '24060124120044', 'Hamsterball Rolling Game',   'https://youtu.be/X7-EdNZdfr8'),
    (5,  'Reynaldi Bertinus Hutagaol',             '24060124140157', '', ''),
    (5,  'Mohammad Banyuputra Eka Pramuditha',     '24060124140193', '', ''),
    (5,  'Ida Bagus Ngurah Yudistira Kemenuh',     '24060124140128', '', ''),
    (6,  'Syifa Aeni Mudrikah',                    '24060124120043', 'Love Simulator Game',        'https://youtu.be/OxwGDezCivo'),
    (6,  'Galang Bintang Ramadhan',                '24060124130101', '', ''),
    (6,  'Puti Shasta Khafiyani',                  '24060124140132', '', ''),
    (6,  'Revalina Salwa Aliya Wicaksono Prabowo', '24060124140155', '', ''),
    (7,  'Ananda Bagus Tri Utomo',                 '24060122130091', 'Truck Simulator',            'https://youtu.be/rWW6hoEYtHc'),
    (7,  'Muhamad Hafidz Zulfikar',                '24060122140141', '', ''),
    (7,  'Nabil Razaki Herman',                    '24060122140147', '', ''),
    (7,  'Rafi Deandra',                           '24060122140122', '', ''),
    (8,  'Raffa Putra Nugroho',                    '24060124130085', '3D Train Simulator',         'https://youtu.be/EqYGuwdQ4b8'),
    (8,  'Laurensius Brian Prayoga',               '24060124130077', '', ''),
    (8,  'Iza Yunus Andhika',                      '24060124140153', '', ''),
    (9,  'Nawaal Hanif Mumtaz Arriye',             '24060124120041', 'EZ Flappy Bird Game',        'https://youtu.be/rOyCfo1IjeY'),
    (9,  'Galvin Shalahudin Mumtaz',               '24060124140162', '', ''),
    (9,  'Rizky Saefirdaus',                       '24060124120001', '', ''),
    (9,  'Muhammad Nauval Fadli',                  '24060124120027', '', ''),
    (10, 'Adhyaksa Margandatua Banjar Nahor',      '24060124140152', 'Quiz Maze Escape Game',      'https://youtu.be/VV_jRwETrEY'),
    (10, 'Raaihan Lazuardi',                       '24060124140178', '', ''),
    (10, 'Farhan Muhtarram',                       '24060124140185', '', ''),
    (10, 'Ganendra Satya Sindhunata',              '24060124120025', '', ''),
    (11, 'Michael Stevano',                        '24060124140187', 'Tower Stacker Game',         'https://youtu.be/Q-S8BUXXiSE'),
]


DATA_GTI_C = [
    (1,  '-', '-', '-', '-'),
    (2,  '-', '-', '-', '-'),
    (3,  'Ruth Septriana Sipangkar',           '24060124120024', '-', 'https://youtu.be/3wSXNM-3HzA'),
    (3,  'Sarifa Nuha Ardanti Jusmar',         '24060124130082', '-', ''),
    (3,  'Syafira Azka Ramadhani',             '24060124130088', '-', ''),
    (3,  'Yasmina Syahidah',                   '24060124130116', '-', ''),
    (4,  'Akbar Mukti Wibowo',                 '24060124130063', '-', 'https://youtu.be/dBWuYekfPr0'),
    (4,  'Maulana Ghazzam Adil Al Faiq',       '24060124130083', '-', ''),
    (4,  'Muhammad Izzat Fauzan Putra Arya',   '24060124130096', '-', ''),
    (4,  'Muhammad Rofad Hamdani',             '24060124130117', '-', ''),
    (5,  'Annis Fakhiroh Akbar',               '24060124130110', '-', 'https://youtu.be/_HqhHM9IX80'),
    (5,  'Binar Ridha Wiritanaya',             '24060124140143', '-', ''),
    (5,  'Nabila Kayla Rafa',                  '24060124120022', '-', ''),
    (6,  'Farras Hilmy Zaidan',                '24060124120003', '-', 'https://youtu.be/KcBhLfcCoas'),
    (6,  'Imam Alfarezzel',                    '24060124120028', '-', ''),
    (6,  'Marco Falias Pangkado',              '24060124130112', '-', ''),
    (6,  "Haydar Rafi' Sultansyah",            '24060124120023', '-', ''),
    (7,  'Mohammad Sulthon Ariefin',           '24060124130104', '-', 'https://youtu.be/HQRgEdBYV2k'),
    (7,  'Haikal Rafli Sembiring',             '24060124130079', '-', ''),
    (7,  'Rio Setiawan Hastanu Putra',         '24060124130068', '-', ''),
    (7,  'Naufal Akbar Nugroho',               '24060124130057', '-', ''),
    (8,  'Azka Aqylla Maulana',                '24060124140195', '-', 'https://youtu.be/OrlicuJV3tQ'),
    (8,  'Akmal Kafli Anan',                   '24060124120042', '-', ''),
    (8,  'Adel Rayyan Hakim',                  '24060124140173', '-', ''),
    (8,  'Agil Yudis Wibawa',                  '24060124120045', '-', ''),
    (9,  'Husni Ulyaa Khanifah',               '24060124120021', '-', 'https://youtu.be/ewmR8Pro5bg'),
    (9,  'Dian Berlian Hutasoit',              '24060124120005', '-', ''),
    (9,  'Christianna Olivia Juniarti M',      '24060124140168', '-', ''),
    (9,  'Dian Aulya Dewiyani',                '24060124130059', '-', ''),
    (10, 'Aswalila Adha Putri Telaumbanua',    '24060124120014', '-', '-'),
    (10, 'Alodia Evelyn Pratikno',             '24060124130087', '-', ''),
    (10, 'Arini Latifatul Qalbiah',            '24060124140136', '-', ''),
    (10, 'Aprillia Abel Cleodora',             '24060124140176', '-', ''),
    (11, 'Muhammad Abhista Pratama Sava',      '24060124130058', '-', 'https://youtu.be/IpFz4pn92Ow'),
    (11, 'Wipin Saputra Poh',                  '24060124130080', '-', ''),
    (12, 'Aufaarel Nabiil Aryadh Mecca',       '24060124140206', '-', 'https://youtu.be/rEJk762uNYw'),
    (12, 'Muhammad Zaidan Alfarizi',           '24060124130102', '-', ''),
    (12, 'Shifa Buja Jauza',                   '24060124140182', '-', ''),
]


# GTI D: hanya ada data kelompok + link (nama/NIM belum tersedia di spreadsheet)
DATA_GTI_D = [
    (1,  '-', '-', '-', 'https://youtu.be/u5T9oImjFh8'),
    (2,  '-', '-', '-', 'https://youtu.be/9r7-SSIk6kQ'),
    (3,  '-', '-', '-', 'https://youtu.be/zBAv8qhIzN4'),
    (4,  '-', '-', '-', 'https://youtu.be/3ojOKIl4RUs'),
    (5,  '-', '-', '-', 'https://youtu.be/18sq-LZdQtc'),
    (6,  '-', '-', '-', 'https://youtu.be/EYOkbbIlPjc'),
    (7,  '-', '-', '-', 'https://youtu.be/LGzCdTaglz0'),
    (8,  '-', '-', '-', 'https://youtu.be/kUmYg1qulW4'),
    (9,  '-', '-', '-', 'https://youtu.be/6FlXx3K7jj8'),
    (10, '-', '-', '-', 'https://youtu.be/5uoeRSHvfQI'),
]

DATA_GTI_E = [
    (1,  '-', '-', '-', '-'),
    (2,  'Varissa Nabila Kifli',              '24060124140125', 'Game Maze 3D',       'https://youtu.be/uxiIlfewLVs'),
    (2,  'Wahyu Aji Gumelar Tri Nugroho',     '24060124140134', '', ''),
    (2,  'Claudia Meitania Putri',            '24060124140188', '', ''),
    (2,  'Adelia Clearesta',                  '24060124140204', '', ''),
    (3,  'Farhan Dwiyan Akbar',               '24060124140137', 'Game Runner',        '-'),
    (3,  'Haikal Imam Ridha',                 '24060124130097', '', ''),
    (3,  'Harits Permana',                    '24060124140131', '', ''),
    (3,  'Joshua Satria Kusuma',              '24060124130113', '', ''),
    (4,  'Aron Sorimuda Johanes Pasaribu',    '24060124130086', 'Sistem Tata Surya',  'https://youtu.be/8rbqDYMV0Bo'),
    (4,  'Husein Avicenna',                   '24060124120047', '', ''),
    (4,  'Rahmat Argyandha Aminuddin',        '24060124130061', '', ''),
    (5,  'Gregorius Septiano Ariadi',         '24060124120026', 'Baldis Game',        '-'),
    (5,  'Dhimas Reza Nafi Wahyudi',          '24060124120010', '', ''),
    (5,  'Djuan Setyo Jati',                  '24060124140163', '', ''),
    (5,  'Fadhil Yaafi Widodo',               '24060124140169', '', ''),
    (6,  'Anggita Kirana Puspa',              '24060124130064', 'ZombieVerse',        '-'),
    (6,  'Felicia Evelina',                   '24060124120012', '', ''),
    (6,  'Raffi Arditama',                    '24060124120020', '', ''),
    (6,  'Rafif Setya Imaduddin',             '24060124130115', '', ''),
    (7,  'Nadia Azura Nurhaniya',             '24060124120019', 'Backroom Escape',    '-'),
    (7,  'Muchamad Yuda Tri Ananda',          '24060124110142', '', ''),
    (7,  'Aditya Sultonul Ulya',              '24060124120006', '', ''),
    (7,  'Rayhan Gerard Darmawan',            '24060124140177', '', ''),
    (8,  'Aqiatillah Rezi Zhafran',           '24060124140124', 'FreakyMaze',         '-'),
    (8,  'Muhammad Farhan Abdul Azis',        '24060124140166', '', ''),
    (8,  'Muhammad Fauzan Akbar',             '24060124140139', '', ''),
    (8,  'Quinta Aurabiansyah',               '24060124120016', '', ''),
    (9,  'Akmal Dzaki Rahmatullah',           '24060124140151', 'Hill Climbing',      'https://youtu.be/IVyrxIBSRzE'),
    (9,  'Wahyu Eko Setyo Pribowo',           '24060124120015', '', ''),
    (10, 'Muhammad Zaidaan Ardiyansyah',      '24060124140200', "Ruang's Games",      '-'),
    (10, 'Muhammad Kemal Faza',               '24060124120013', '', ''),
    (10, 'Anintya Abhi Wiryateja',            '24060124130053', '', ''),
    (11, 'Mischa Nathanael Lumban Tobing',    '24060124140175', 'Caterpillar Game 3D','-'),
    (11, 'Levi Ramot Siahaan',                '24060124130067', '', ''),
]


# ============================================================
# RUBRIK & KOLOM PENILAIAN (konsisten semua kelas)
# ============================================================
RUBRIK = [
    # (kode, judul_singkat,             judul_panjang_untuk_tooltip,                      bobot)
    ('A', 'Kualitas Video BTS\n(max 20)',   'Durasi ≤5 menit, judul/deskripsi format lengkap, kualitas visual & audio baik',                 20),
    ('B', 'Representasi\nObjek 3D\n(max 10)', 'Objek 3D divisualisasikan dengan benar (mesh, vertex, edge, face)',                          10),
    ('C', 'Proyeksi\nOrthographic\n(max 10)', 'Proyeksi orthographic diterapkan dan dijelaskan dengan benar',                               10),
    ('D', 'Proyeksi\nPerspektif\n(max 10)',   'Perspektif 1/2/3 point diterapkan dan dijelaskan dengan benar',                              10),
    ('E', 'Geometri\n(max 8)',                'Transformasi geometri (translasi, rotasi, skala) diterapkan',                                  8),
    ('F', 'Kamera\n(max 8)',                  'Setup kamera (FOV, near/far plane, viewport) diterapkan',                                      8),
    ('G', 'Cahaya\n(max 8)',                  'Pencahayaan (ambient, diffuse, specular) diterapkan',                                          8),
    ('H', 'Karakteristik\nPermukaan\n(max 8)','Material/tekstur (warna, roughness, metallic, dll.) diterapkan',                              8),
    ('I', 'Algoritma\nRendering\n(max 8)',    'Algoritma rendering (wireframe/rasterization/raytracing) dijelaskan',                          8),
    ('J', 'Shading\n(max 10)',                'Model shading (flat/Gouraud/Phong) diterapkan dan dijelaskan',                               10),
]
# Total bobot = 100

HEADERS = ['No. Kelompok', 'NIM', 'Nama Mahasiswa', 'Nama Proyek', 'Link Video'] \
          + [r[1] for r in RUBRIK] \
          + ['TOTAL\n(max 100)', 'Grade', 'Catatan']

# ============================================================
# WARNA
# ============================================================
COLOR_HEADER_BG   = 'FF1F4E79'  # biru gelap
COLOR_HEADER_FONT = 'FFFFFFFF'  # putih
COLOR_RUBRIK_BG   = 'FF2E75B6'  # biru menengah
COLOR_SUB_HEADER  = 'FFD6E4F0'  # biru muda
COLOR_ROW_ODD     = 'FFEEF4FB'  # strip biru sangat muda
COLOR_ROW_EVEN    = 'FFFFFFFF'  # putih
COLOR_TOTAL_BG    = 'FFFFF2CC'  # kuning muda
COLOR_LINK        = 'FF0563C1'  # biru link
COLOR_WARN        = 'FFFFC7CE'  # merah muda (belum kumpul)
COLOR_GRADE_BG    = 'FFE2EFDA'  # hijau muda

GRADE_SCALE = [
    (85, 'A'),
    (75, 'AB'),
    (65, 'B'),
    (55, 'BC'),
    (45, 'C'),
    (35, 'D'),
    (0,  'E'),
]


# ============================================================
# HELPER FUNGSI STYLING
# ============================================================
def thin_border():
    s = Side(style='thin', color='FF000000')
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style='medium', color='FF000000')
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, color='FF000000', size=10):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def center_align(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def style_cell(cell, bg=None, bold=False, color='FF000000', size=10,
               align='center', border=True, wrap=True):
    if bg:
        cell.fill = fill(bg)
    cell.font = font(bold=bold, color=color, size=size)
    cell.alignment = center_align(wrap) if align == 'center' else left_align(wrap)
    if border:
        cell.border = thin_border()

def grade_from_total(total):
    if not isinstance(total, (int, float)):
        return ''
    for threshold, grade in GRADE_SCALE:
        if total >= threshold:
            return grade
    return 'E'


# ============================================================
# BARIS RUBRIK (baris ke-2 setelah header, untuk penjelasan bobot)
# ============================================================
def write_rubrik_row(ws, rubrik_row_idx):
    """Tulis baris keterangan rubrik di bawah header."""
    keterangan = ['', '', '', '', ''] + [r[3] for r in RUBRIK] + ['', '', '']
    for col_idx, val in enumerate(keterangan, start=1):
        cell = ws.cell(row=rubrik_row_idx, column=col_idx)
        cell.value = str(val) if val else ''
        style_cell(cell, bg='FFD9E1F2', bold=False, color='FF1F3864',
                   size=8, align='center', border=True)

# ============================================================
# FUNGSI UTAMA: BUAT SATU SHEET
# ============================================================
def build_sheet(ws, kelas_name, data):
    # ---- Freeze panes ----
    ws.freeze_panes = 'A4'

    # ---- Baris 1: judul kelas ----
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f'PENILAIAN TUGAS BESAR GTI {kelas_name} — SEMESTER GENAP 2025/2026'
    style_cell(title_cell, bg=COLOR_HEADER_BG, bold=True,
               color=COLOR_HEADER_FONT, size=13, align='center')

    # ---- Baris 2: header kolom ----
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = h
        # Kolom nilai pakai warna berbeda
        if col_idx <= 5:
            bg = 'FF1F4E79'
        elif col_idx == len(HEADERS) - 2:   # TOTAL
            bg = 'FFBF8F00'
        elif col_idx >= len(HEADERS) - 1:   # Grade, Catatan
            bg = 'FF375623'
        else:
            bg = COLOR_RUBRIK_BG
        style_cell(cell, bg=bg, bold=True,
                   color='FFFFFFFF', size=9, align='center')

    # ---- Baris 3: keterangan rubrik ----
    write_rubrik_row(ws, 3)

    # ---- Lebar kolom ----
    col_widths = [12, 18, 30, 22, 42] + [13] * len(RUBRIK) + [12, 8, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Tinggi baris ----
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 52
    ws.row_dimensions[3].height = 30

    # ---- Data mahasiswa ----
    prev_kelompok = None
    row_colors = [COLOR_ROW_ODD, COLOR_ROW_EVEN]
    color_toggle = 0

    for data_row_idx, (klp, nama, nim, proyek, link) in enumerate(data):
        row_num = data_row_idx + 4  # mulai dari baris 4

        if klp != prev_kelompok:
            color_toggle = 1 - color_toggle
            prev_kelompok = klp

        row_bg = row_colors[color_toggle]
        is_empty = (nama == '-' or nama == '')

        # ---- Kolom identitas ----
        vals = [klp, nim, nama, proyek if proyek else '', link if link else '']
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_num, column=ci)
            cell.value = v
            bg = COLOR_WARN if is_empty else row_bg
            cell.fill = fill(bg)
            cell.font = Font(name='Calibri', size=10,
                             color=COLOR_LINK if (ci == 5 and v.startswith('http')) else 'FF000000')
            cell.alignment = center_align() if ci in (1,) else left_align()
            cell.border = thin_border()
            if ci == 5 and v.startswith('http'):
                ws.cell(row=row_num, column=ci).hyperlink = v

        # ---- Kolom nilai (kosong, siap diisi) ----
        nilai_start_col = 6
        total_col       = nilai_start_col + len(RUBRIK)
        grade_col       = total_col + 1
        catatan_col     = total_col + 2

        for ci in range(nilai_start_col, total_col):
            cell = ws.cell(row=row_num, column=ci)
            cell.value = None
            cell.fill  = fill(COLOR_WARN if is_empty else row_bg)
            cell.font  = Font(name='Calibri', size=10)
            cell.alignment = center_align()
            cell.border = thin_border()

        # ---- Kolom TOTAL (formula SUM) ----
        first_val_col = get_column_letter(nilai_start_col)
        last_val_col  = get_column_letter(total_col - 1)
        total_cell = ws.cell(row=row_num, column=total_col)
        if not is_empty:
            total_cell.value = f'=SUM({first_val_col}{row_num}:{last_val_col}{row_num})'
        else:
            total_cell.value = None
        style_cell(total_cell, bg=COLOR_TOTAL_BG if not is_empty else COLOR_WARN,
                   bold=True, size=10, align='center')

        # ---- Kolom Grade (formula IF) ----
        tc = get_column_letter(total_col)
        grade_cell = ws.cell(row=row_num, column=grade_col)
        if not is_empty:
            grade_cell.value = (
                f'=IF({tc}{row_num}="","",IF({tc}{row_num}>=85,"A",'
                f'IF({tc}{row_num}>=75,"AB",IF({tc}{row_num}>=65,"B",'
                f'IF({tc}{row_num}>=55,"BC",IF({tc}{row_num}>=45,"C",'
                f'IF({tc}{row_num}>=35,"D","E")))))))'
            )
        else:
            grade_cell.value = None
        style_cell(grade_cell, bg=COLOR_GRADE_BG if not is_empty else COLOR_WARN,
                   bold=True, size=10, align='center')

        # ---- Kolom Catatan ----
        cat_cell = ws.cell(row=row_num, column=catatan_col)
        if is_empty:
            cat_cell.value = 'Belum mengumpulkan / data tidak tersedia'
        else:
            cat_cell.value = ''
        style_cell(cat_cell, bg=COLOR_WARN if is_empty else row_bg,
                   size=9, align='left')

        ws.row_dimensions[row_num].height = 18

    # ---- Baris ringkasan di akhir ----
    last_data_row = len(data) + 3
    summary_row   = last_data_row + 2
    ws.merge_cells(start_row=summary_row, start_column=1,
                   end_row=summary_row, end_column=5)
    s_cell = ws.cell(row=summary_row, column=1)
    s_cell.value = 'SKALA NILAI: A ≥ 85 | AB 75–84 | B 65–74 | BC 55–64 | C 45–54 | D 35–44 | E < 35'
    style_cell(s_cell, bg='FFFCE4D6', bold=True, color='FF833C00',
               size=9, align='left')


# ============================================================
# SHEET RUBRIK LENGKAP
# ============================================================
def build_rubrik_sheet(ws):
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 55

    # Judul
    ws.merge_cells('A1:E1')
    t = ws.cell(row=1, column=1)
    t.value = 'RUBRIK PENILAIAN TUGAS BESAR — GRAFIKA & TEKNOLOGI INFORMASI (GTI) 2026'
    style_cell(t, bg=COLOR_HEADER_BG, bold=True, color='FFFFFFFF', size=13)

    # Sub-judul
    ws.merge_cells('A2:E2')
    s = ws.cell(row=2, column=1)
    s.value = 'Tugas: Video Behind The Scenes (BTS) — Maks 5 Menit — Upload YouTube'
    style_cell(s, bg=COLOR_RUBRIK_BG, bold=True, color='FFFFFFFF', size=11)

    # Header tabel rubrik
    headers_r = ['Kode', 'Komponen', 'Indikator / Deskripsi', 'Bobot (Max)', 'Kriteria Penilaian Detail']
    for ci, h in enumerate(headers_r, start=1):
        cell = ws.cell(row=3, column=ci)
        cell.value = h
        style_cell(cell, bg='FF2E75B6', bold=True, color='FFFFFFFF', size=10)
    ws.row_dimensions[3].height = 18

    # Detail rubrik
    rubrik_detail = [
        ('A', 'Kualitas Video BTS', 'Durasi ≤5 menit, judul format [Nama Projek] BTS ([Nama Panggilan Ketua]), deskripsi berisi nama+NIM semua anggota & cerita singkat, kualitas visual & audio baik', 20,
         '18–20: Semua terpenuhi dengan sangat baik\n14–17: Sebagian besar terpenuhi, minor kekurangan\n10–13: Beberapa komponen kurang\n<10: Banyak komponen tidak terpenuhi'),
        ('B', 'Representasi Objek 3D', 'Teknik representasi objek 3D diterapkan dengan benar (mesh, vertex, edge, face, polygon)', 10,
         '9–10: Representasi 3D lengkap & jelas ditunjukkan\n7–8: Representasi 3D ditunjukkan, sedikit kurang detail\n4–6: Ditunjukkan tapi tidak dijelaskan\n1–3: Tidak diterapkan / sangat kurang'),
        ('C', 'Proyeksi Orthographic', 'Proyeksi orthographic (tampak atas/depan/samping) diterapkan dan dijelaskan secara eksplisit dalam video', 10,
         '9–10: Diterapkan dan dijelaskan dengan sangat jelas\n7–8: Diterapkan, penjelasan cukup\n4–6: Ada tapi penjelasan kurang\n1–3: Tidak ada / sangat kurang'),
        ('D', 'Proyeksi Perspektif', 'Perspektif 1-point, 2-point, atau 3-point diterapkan dan dijelaskan. Poin ekstra jika lebih dari 1 jenis.', 10,
         '9–10: Diterapkan & dijelaskan dengan sangat jelas\n7–8: Diterapkan, penjelasan cukup\n4–6: Ada tapi penjelasan kurang\n1–3: Tidak ada / sangat kurang'),
        ('E', 'Geometri (Rendering Pipeline)', 'Transformasi geometri (translasi, rotasi, skala, shear) diterapkan pada objek 3D dalam proyek', 8,
         '7–8: Transformasi lengkap & terlihat jelas\n5–6: Transformasi dasar ada\n3–4: Sebagian ada\n1–2: Tidak ada / tidak jelas'),
        ('F', 'Kamera (Rendering Pipeline)', 'Setup kamera yang benar: FOV, near/far clipping plane, viewport, posisi/orientasi kamera', 8,
         '7–8: Setup kamera lengkap & dijelaskan\n5–6: Kamera di-set tapi sebagian kurang\n3–4: Ada kamera tapi tidak dijelaskan\n1–2: Tidak ada'),
        ('G', 'Cahaya (Rendering Pipeline)', 'Pencahayaan: ambient, diffuse, specular — minimal 2 jenis diterapkan dan terlihat dalam render', 8,
         '7–8: Minimal 2 jenis cahaya, terlihat & dijelaskan\n5–6: 1 jenis cahaya, cukup\n3–4: Ada cahaya tapi tidak dijelaskan\n1–2: Tidak ada'),
        ('H', 'Karakteristik Permukaan', 'Material / tekstur objek: warna, roughness, metallic, atau tekstur gambar diterapkan', 8,
         '7–8: Material kaya & beragam per objek\n5–6: Material dasar ada\n3–4: Warna polos saja\n1–2: Tidak ada material'),
        ('I', 'Algoritma Rendering', 'Algoritma rendering dijelaskan: wireframe, scanline rasterization, ray tracing, atau lainnya', 8,
         '7–8: Algoritma dijelaskan & terlihat dalam video\n5–6: Disebutkan tapi kurang detail\n3–4: Tersirat tapi tidak eksplisit\n1–2: Tidak dijelaskan'),
        ('J', 'Shading', 'Model shading diterapkan dan dijelaskan: flat shading, Gouraud shading, Phong shading', 10,
         '9–10: Shading tepat, terlihat jelas, dijelaskan detail\n7–8: Shading ada, penjelasan cukup\n4–6: Shading ada tapi tidak dijelaskan\n1–3: Tidak ada / tidak jelas'),
    ]

    for ri, (kode, komponen, indikator, bobot, kriteria) in enumerate(rubrik_detail, start=4):
        row_bg = 'FFEEF4FB' if ri % 2 == 0 else 'FFFFFFFF'
        vals = [kode, komponen, indikator, bobot, kriteria]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = val
            cell.fill = fill(COLOR_TOTAL_BG if ci == 4 else row_bg)
            cell.font = Font(name='Calibri', size=9,
                             bold=(ci in (1, 4)),
                             color='FF833C00' if ci == 4 else 'FF000000')
            cell.alignment = center_align() if ci in (1, 4) else left_align()
            cell.border = thin_border()
        ws.row_dimensions[ri].height = 70

    # Total bobot
    total_row = len(rubrik_detail) + 4
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    tot = ws.cell(row=total_row, column=1)
    tot.value = 'TOTAL BOBOT'
    style_cell(tot, bg=COLOR_HEADER_BG, bold=True, color='FFFFFFFF', size=11)
    tot_val = ws.cell(row=total_row, column=4)
    tot_val.value = 100
    style_cell(tot_val, bg='FFBF8F00', bold=True, color='FFFFFFFF', size=12)
    ws.row_dimensions[total_row].height = 22


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()

    # Hapus sheet default
    wb.remove(wb.active)

    # Sheet RUBRIK di posisi pertama
    ws_rubrik = wb.create_sheet('RUBRIK', 0)
    build_rubrik_sheet(ws_rubrik)

    # Sheet tiap kelas
    kelas_data = [
        ('GTI A', DATA_GTI_A),
        ('GTI B', DATA_GTI_B),
        ('GTI C', DATA_GTI_C),
        ('GTI D', DATA_GTI_D),
        ('GTI E', DATA_GTI_E),
    ]
    for sheet_name, data in kelas_data:
        ws = wb.create_sheet(sheet_name)
        build_sheet(ws, sheet_name.split()[-1], data)
        print(f'  ✓ Sheet {sheet_name} selesai ({len(data)} baris data)')

    out_path = '/projects/sandbox/sonet-product/nilai-nilai/gti-2026/Penilaian_GTI_2026.xlsx'
    wb.save(out_path)
    print(f'\n✅ File Excel berhasil dibuat: {out_path}')

if __name__ == '__main__':
    main()
