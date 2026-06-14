"""
Generate PENILAIAN-PPT-2.xlsx
IoT Mat-D 2026 — Penilaian Presentasi Kelompok PPT-2

Rubrik (5 kriteria, total 100):
  P1  Arsitektur IoT              25
  P2  Alur Data & Protokol        20
  P3  Organisasi & Manajemen IoT  20
  P4  Bisnis IoT                  20
  P5  Kualitas Presentasi         15
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────────
# DATA KELOMPOK
# Format: (No, Nama_Kelompok, Topik_Proyek, Anggota)
# Anggota diisi berdasarkan inferensi proyek dari data Kuis.
# Dosen/asisten dapat menyesuaikan nama kelompok & anggota.
# ─────────────────────────────────────────────────────────────────
groups = [
    (1,  "Kelompok 1",  "Smart Gas Leak Detector",
         "Khoiriyatun · Intan Tama Jessica Purba · Imelda Nuris Syifa"),
    (2,  "Kelompok 2",  "EWS Banjir Rob",
         "Hentrika Aji Pamungkas · Evlyna Fedora Hartanto · Nadia Bilqis"),
    (3,  "Kelompok 3",  "Water Level & Smart Water Monitoring",
         "Agnia Faradha Listya · Lusiana Kezia Arla Gracyani · Dini Permata Anisa · Alya Kamila Lubna"),
    (4,  "Kelompok 4",  "Deteksi Rob & Alarm Banjir",
         "Muhammad Fadel Arnanda · Ramayana Sugiyo Pranoto · Bayhaqiy Ahmad"),
    (5,  "Kelompok 5",  "Absensi Pintar / Presensi IoT",
         "Abdulloh Ibnu Musa · Nabila Fiesta Ramadhani · Rizki Fathiya Nur Khairunissa"),
    (6,  "Kelompok 6",  "AEWG — Accident Early Warning Grid",
         "Edria Filda Tsana · Nimatul Karimah · Sadira Najla Filzah Aryamanto"),
    (7,  "Kelompok 7",  "Odor / Wound Detector IoT",
         "Ahmad Zidane Ainul Yaqin · Ezra Aryasatya · Achmad Fachreza Aryadewa · Dafa Septa Ramadhani"),
    (8,  "Kelompok 8",  "Railguard — Smart Railway Safety",
         "Ryan Anggit Nugroho · Alwan Rasyid Ramadhan · Muhammad Haykal Harsya Mevki"),
    (9,  "Kelompok 9",  "Smart Waste / TPS Pintar",
         "Alif Alfiyansah · Shafana Puja Pitaloka · Ihsan Izzat Ibrahim · Lestari Kopipah Mandasari"),
    (10, "Kelompok 10", "TPS Smart Sampah & Deteksi Sampah",
         "Mahira Layina Raisha Azhar · Diah Pertiwi Difa Nur Ilmaini · Rahma Nabila Ramadhani · Nadyatul Aulia · Amanda Marchelia C. Putri"),
    (11, "Kelompok 11", "PantauKu — Smart Wearable & Child Safety",
         "Wiwi Lindawati Misnadin · Syifa Ayudya Nurhafiza · Cantika Putri Maharani"),
    (12, "Kelompok 12", "AeroSentry — Pemantauan Kualitas Udara",
         "Orva Luttayya Ananditasuntoro · Rieska Annisa Salsabilla · Abyan Hilmy Fikri Sagala"),
    (13, "Kelompok 13", "Smart Air & Pollution Monitoring",
         "Fanny Laviqnia Lova · Bagoes Satria Jagad Dhita · Zaskia Hayatunufus · Muhammad Najmi Fathoni · Defan Atara Fahrezy · Hazel Ihsan Fadillah · Ananda Setiawan · Dira Artafirasha"),
]

# Bobot maksimal tiap kriteria
BOBOT = {"P1": 25, "P2": 20, "P3": 20, "P4": 20, "P5": 15}

# ─────────────────────────────────────────────────────────────────
# RUBRIK DATA (untuk Sheet 2)
# ─────────────────────────────────────────────────────────────────
rubrik_rows = [
    ("P1", "Arsitektur IoT", 25,
     "Ketepatan layer (Perception→Network→Application / 3–5 layer), "
     "pemilihan hardware (sensor, MCU, gateway, cloud), kesesuaian komponen dengan proyek.",
     "22–25", "19–21", "15–18", "10–14", "0–9"),
    ("P2", "Alur Data & Protokol", 20,
     "Kejelasan alur data (input→proses→output), protokol komunikasi (MQTT/HTTP/LoRa/dsb.), "
     "integrasi antar-layer, dan justifikasi pilihan protokol.",
     "18–20", "15–17", "12–14", "8–11", "0–7"),
    ("P3", "Organisasi & Manajemen IoT", 20,
     "Peran stakeholder (pengguna, operator, mitra), tata kelola data "
     "(kepemilikan, akses, tanggung jawab), pembagian tugas tim.",
     "18–20", "15–17", "12–14", "8–11", "0–7"),
    ("P4", "Bisnis IoT", 20,
     "Target pasar spesifik, model bisnis / revenue stream, "
     "minimal 2 aspek TLBMC (Ekonomi, Lingkungan, Sosial), value proposition.",
     "18–20", "15–17", "12–14", "8–11", "0–7"),
    ("P5", "Kualitas Presentasi", 15,
     "Slide terstruktur & terbaca, penyampaian lancar, "
     "mampu menjawab pertanyaan, ketepatan waktu.",
     "13–15", "11–12", "9–10", "6–8", "0–5"),
]

# ─────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────
def thin_border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_outer(color="2E75B6"):
    m = Side(style="medium", color=color)
    return Border(left=m, right=m, top=m, bottom=m)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def grade_fill(huruf):
    m = {"A":"C6EFCE","AB":"DBEFCC","B":"FFEB9C","BC":"FFD966",
         "C":"FCE4D6","CD":"F4CCCC","D":"EA9999","E":"CC0000"}
    return m.get(huruf, "FFFFFF")

def grade_font_color(huruf):
    return "FFFFFF" if huruf == "E" else "000000"

def huruf_from_total(total):
    if total >= 80: return "A"
    elif total >= 75: return "AB"
    elif total >= 70: return "B"
    elif total >= 65: return "BC"
    elif total >= 60: return "C"
    elif total >= 55: return "CD"
    elif total >= 45: return "D"
    else: return "E"

# ─────────────────────────────────────────────────────────────────
# SHEET 1: PENILAIAN
# ─────────────────────────────────────────────────────────────────
def build_penilaian_sheet(ws):
    # Column layout (1-indexed):
    # A=No  B=Nama Kelompok  C=Topik Proyek  D=Anggota
    # E=P1  F=P2  G=P3  H=P4  I=P5
    # J=Total  K=Huruf  L=Komentar / Kekurangan

    COL = {"No":1,"Kelompok":2,"Topik":3,"Anggota":4,
           "P1":5,"P2":6,"P3":7,"P4":8,"P5":9,
           "Total":10,"Huruf":11,"Komentar":12}
    N_COLS = 12

    # ── Row 1: Main title ──────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    c = ws.cell(row=1, column=1,
                value="PENILAIAN PPT-2 — PRESENTASI KELOMPOK  |  IoT Mat-D 2026")
    c.font      = font(bold=True, size=14, color="FFFFFF")
    c.fill      = fill("1F4E79")
    c.alignment = align("center","center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: subtitle + bobot ───────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
    c = ws.cell(row=2, column=1,
                value="Rubrik: P1 Arsitektur(25) · P2 Alur Data(20) · P3 Organisasi(20) · P4 Bisnis(20) · P5 Presentasi(15)  |  A≥80 · AB 75–79 · B 70–74 · BC 65–69 · C 60–64 · CD 55–59 · D 45–54 · E<45")
    c.font      = font(italic=True, size=9, color="444444")
    c.fill      = fill("D6E4F0")
    c.alignment = align("center","center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: bobot sub-headers ──────────────────────────────
    bobot_labels = {
        COL["P1"]: "P1\n(maks 25)",
        COL["P2"]: "P2\n(maks 20)",
        COL["P3"]: "P3\n(maks 20)",
        COL["P4"]: "P4\n(maks 20)",
        COL["P5"]: "P5\n(maks 15)",
    }
    for col in range(1, N_COLS+1):
        c = ws.cell(row=3, column=col)
        if col in bobot_labels:
            c.value = bobot_labels[col]
        ws.cell(row=3, column=col).fill = fill("EBF3FB")
    ws.row_dimensions[3].height = 4

    # ── Row 4: column headers ─────────────────────────────────
    headers = {
        COL["No"]:       "No",
        COL["Kelompok"]: "Nama Kelompok",
        COL["Topik"]:    "Topik Proyek",
        COL["Anggota"]:  "Anggota",
        COL["P1"]:       "P1\nArsitektur IoT",
        COL["P2"]:       "P2\nAlur Data & Protokol",
        COL["P3"]:       "P3\nOrganisasi IoT",
        COL["P4"]:       "P4\nBisnis IoT",
        COL["P5"]:       "P5\nKualitas Presentasi",
        COL["Total"]:    "Total",
        COL["Huruf"]:    "Huruf",
        COL["Komentar"]: "Komentar & Kekurangan",
    }
    for col, label in headers.items():
        c = ws.cell(row=4, column=col, value=label)
        c.font      = font(bold=True, size=10, color="FFFFFF")
        c.fill      = fill("2E75B6")
        c.alignment = align("center","center", wrap=True)
        c.border    = thin_border("1F4E79")
    ws.row_dimensions[4].height = 36

    # ── Data rows ─────────────────────────────────────────────
    DATA_START = 5
    bdr = thin_border()
    for idx, (no, nama, topik, anggota) in enumerate(groups):
        r = DATA_START + idx
        row_bg = "F7FBFF" if idx % 2 == 0 else "FFFFFF"
        alt = fill(row_bg)

        # No
        c = ws.cell(row=r, column=COL["No"], value=no)
        c.font = font(bold=True, size=10); c.fill = fill("D6E4F0")
        c.alignment = align("center","center"); c.border = bdr

        # Nama Kelompok
        c = ws.cell(row=r, column=COL["Kelompok"], value=nama)
        c.font = font(bold=True, size=10); c.fill = fill("EBF3FB")
        c.alignment = align("left","center", wrap=True); c.border = bdr

        # Topik Proyek
        c = ws.cell(row=r, column=COL["Topik"], value=topik)
        c.font = font(size=10); c.fill = alt
        c.alignment = align("left","center", wrap=True); c.border = bdr

        # Anggota
        c = ws.cell(row=r, column=COL["Anggota"], value=anggota)
        c.font = font(size=9, italic=True); c.fill = alt
        c.alignment = align("left","top", wrap=True); c.border = bdr

        # Score cells P1–P5 (empty, user fills in)
        score_cols = [COL["P1"], COL["P2"], COL["P3"], COL["P4"], COL["P5"]]
        score_letters = [get_column_letter(col) for col in score_cols]
        for ci in score_cols:
            c = ws.cell(row=r, column=ci, value=None)
            c.font = font(size=11, bold=True, color="1F4E79")
            c.fill = fill("FDFEFE")
            c.alignment = align("center","center")
            c.border = thin_border("2E75B6")
            c.number_format = "0"

        # Total (formula)
        total_addr = "+".join(f"{ltr}{r}" for ltr in score_letters)
        total_col_ltr = get_column_letter(COL["Total"])
        c = ws.cell(row=r, column=COL["Total"], value=f"={total_addr}")
        c.font = font(bold=True, size=11, color="1F4E79")
        c.fill = fill("D6E4F0")
        c.alignment = align("center","center")
        c.border = thin_border("2E75B6")
        c.number_format = "0"

        # Huruf (nested IF formula)
        tc = f"{total_col_ltr}{r}"
        huruf_formula = (
            f'=IF({tc}="","",IF({tc}>=80,"A",IF({tc}>=75,"AB",IF({tc}>=70,"B",'
            f'IF({tc}>=65,"BC",IF({tc}>=60,"C",IF({tc}>=55,"CD",IF({tc}>=45,"D","E"))))))))'
        )
        c = ws.cell(row=r, column=COL["Huruf"], value=huruf_formula)
        c.font = font(bold=True, size=11)
        c.fill = fill("EBF3FB")
        c.alignment = align("center","center")
        c.border = thin_border("2E75B6")

        # Komentar (empty)
        c = ws.cell(row=r, column=COL["Komentar"], value="")
        c.font = font(size=10, italic=True, color="444444")
        c.fill = fill("FEFEFE")
        c.alignment = align("left","top", wrap=True)
        c.border = bdr

        ws.row_dimensions[r].height = 60

    # ── Summary row ───────────────────────────────────────────
    SUMROW = DATA_START + len(groups)
    ws.merge_cells(start_row=SUMROW, start_column=1, end_row=SUMROW, end_column=COL["P1"]-1)
    c = ws.cell(row=SUMROW, column=1, value="Rata-rata Kelas")
    c.font = font(bold=True, size=10, color="FFFFFF")
    c.fill = fill("1F4E79"); c.alignment = align("center","center")
    for col in range(1, COL["P1"]): ws.cell(row=SUMROW, column=col).fill = fill("1F4E79")

    for ci, key in zip(
        [COL["P1"], COL["P2"], COL["P3"], COL["P4"], COL["P5"], COL["Total"]],
        ["P1", "P2", "P3", "P4", "P5", "Total"]
    ):
        ltr = get_column_letter(ci)
        rng = f"{ltr}{DATA_START}:{ltr}{DATA_START+len(groups)-1}"
        c = ws.cell(row=SUMROW, column=ci,
                    value=f'=IFERROR(AVERAGEIF({rng},"<>",{rng}),"—")')
        c.font = font(bold=True, size=10, color="FFFFFF")
        c.fill = fill("2E75B6")
        c.alignment = align("center","center")
        c.border = thin_border("1F4E79")
        c.number_format = "0.0"
    ws.row_dimensions[SUMROW].height = 22

    # ── Column widths ─────────────────────────────────────────
    widths = {
        COL["No"]: 4, COL["Kelompok"]: 18, COL["Topik"]: 30, COL["Anggota"]: 45,
        COL["P1"]: 9, COL["P2"]: 9, COL["P3"]: 9, COL["P4"]: 9, COL["P5"]: 9,
        COL["Total"]: 8, COL["Huruf"]: 7, COL["Komentar"]: 60,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Freeze + autofilter ───────────────────────────────────
    ws.freeze_panes = "E5"
    ws.auto_filter.ref = f"A4:{get_column_letter(N_COLS)}4"

    # ── Conditional formatting note: colour Huruf cell ────────
    # (done via static formula above; dynamic coloring requires openpyxl ConditionalFormatting)
    # Add a note row at the top explaining colour coding
    ws.sheet_view.zoomScale = 90


# ─────────────────────────────────────────────────────────────────
# SHEET 2: RUBRIK REFERENSI
# ─────────────────────────────────────────────────────────────────
def build_rubrik_sheet(ws):
    # Title
    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value="RUBRIK PENILAIAN PPT-2  —  IoT Mat-D 2026")
    c.font      = font(bold=True, size=13, color="FFFFFF")
    c.fill      = fill("1F4E79")
    c.alignment = align("center","center")
    ws.row_dimensions[1].height = 26

    # Header row
    rubrik_headers = ["Kode","Kriteria","Bobot","Deskripsi",
                      "Sangat Baik\n(88–100%)","Baik\n(75–87%)",
                      "Cukup\n(60–74%)","Kurang\n(40–59%)",
                      "Sangat Kurang\n(0–39%)"]
    HDR_FILLS = ["2E75B6"] * len(rubrik_headers)
    for ci, (h, hf) in enumerate(zip(rubrik_headers, HDR_FILLS), start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = font(bold=True, size=10, color="FFFFFF")
        c.fill      = fill(hf)
        c.alignment = align("center","center", wrap=True)
        c.border    = thin_border("1F4E79")
    ws.row_dimensions[2].height = 36

    # Rubrik data rows
    row_fills = ["EBF3FB", "FFFFFF"]
    for idx, (kode, nama, bobot, desk, sb, b, c_, k, sk) in enumerate(rubrik_rows):
        r = idx + 3
        rf = fill(row_fills[idx % 2])
        vals = [kode, nama, bobot, desk, sb, b, c_, k, sk]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border    = thin_border()
            cell.alignment = align("center","top", wrap=True) if ci != 4 else align("left","top", wrap=True)
            cell.font      = font(bold=(ci<=2), size=10)
            cell.fill      = rf
        ws.row_dimensions[r].height = 55

    # Huruf scale table
    SROW = len(rubrik_rows) + 4
    ws.merge_cells(start_row=SROW, start_column=1, end_row=SROW, end_column=9)
    c = ws.cell(row=SROW, column=1, value="Skala Huruf (Standar Undip)")
    c.font = font(bold=True, size=11, color="FFFFFF")
    c.fill = fill("2E75B6"); c.alignment = align("center","center")
    ws.row_dimensions[SROW].height = 20

    scale = [
        ("A","80–100","Sangat Baik","C6EFCE","000000"),
        ("AB","75–79","Baik Sekali","DBEFCC","000000"),
        ("B","70–74","Baik","FFEB9C","000000"),
        ("BC","65–69","Cukup Baik","FFD966","000000"),
        ("C","60–64","Cukup","FCE4D6","000000"),
        ("CD","55–59","Kurang Cukup","F4CCCC","000000"),
        ("D","45–54","Kurang","EA9999","000000"),
        ("E","<45","Sangat Kurang","CC0000","FFFFFF"),
    ]
    for si, (h, r_val, ket, bg, fg) in enumerate(scale):
        sr = SROW + 1 + si
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=3)
        c = ws.cell(row=sr, column=1, value=h)
        c.font = font(bold=True, size=11, color=fg); c.fill = fill(bg)
        c.alignment = align("center","center"); c.border = thin_border()
        ws.merge_cells(start_row=sr, start_column=4, end_row=sr, end_column=5)
        c = ws.cell(row=sr, column=4, value=r_val)
        c.font = font(size=10); c.fill = fill(bg)
        c.alignment = align("center","center"); c.border = thin_border()
        ws.merge_cells(start_row=sr, start_column=6, end_row=sr, end_column=9)
        c = ws.cell(row=sr, column=6, value=ket)
        c.font = font(size=10); c.fill = fill(bg)
        c.alignment = align("left","center"); c.border = thin_border()
        ws.row_dimensions[sr].height = 18

    # Column widths for rubrik sheet
    rub_widths = [6, 22, 7, 50, 14, 14, 14, 14, 14]
    for ci, w in enumerate(rub_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 95


# ─────────────────────────────────────────────────────────────────
# SHEET 3: DAFTAR MAHASISWA (referensi crosscheck)
# ─────────────────────────────────────────────────────────────────
students_ref = [
    (1,  "...120004", "Khoiriyatun",                  1),
    (2,  "...120010", "Intan Tama Jessica Purba",      1),
    (3,  "...120016", "Imelda Nuris Syifa",            1),
    (4,  "...120020", "Hentrika Aji Pamungkas",        2),
    (5,  "...120022", "Evlyna Fedora Hartanto",        2),
    (6,  "...120028", "Nadia Bilqis",                  2),
    (7,  "...120032", "Agnia Faradha Listya",          3),
    (8,  "...120036", "Lusiana Kezia Arla Gracyani",   3),
    (9,  "...120040", "Muhammad Fadel Arnanda",        4),
    (10, "...120046", "Ramayana Sugiyo Pranoto",       4),
    (11, "...120056", "Abdulloh Ibnu Musa",            5),
    (12, "...120060", "Nabila Fiesta Ramadhani",       5),
    (13, "...120062", "Edria Filda Tsana",             6),
    (14, "...130064", "Ahmad Zidane Ainul Yaqin",      7),
    (15, "...130076", "Ezra Aryasatya",                7),
    (16, "...130078", "Bayhaqiy Ahmad",                4),
    (17, "...130080", "Dini Permata Anisa",            3),
    (18, "...130082", "Achmad Fachreza Aryadewa",      7),
    (19, "...130084", "Nimatul Karimah",               6),
    (20, "...130097", "Dafa Septa Ramadhani",          7),
    (21, "...130100", "Rizki Fathiya Nur Khairunissa", 5),
    (22, "...130106", "Alya Kamila Lubna",             3),
    (23, "...130110", "Ryan Anggit Nugroho",           8),
    (24, "...130112", "Sadira Najla Filzah Aryamanto", 6),
    (25, "...130116", "Alwan Rasyid Ramadhan",         8),
    (26, "...130122", "Ihsan Izzat Ibrahim",           9),
    (27, "...130126", "Alif Alfiyansah",               9),
    (28, "...130128", "Shafana Puja Pitaloka",         9),
    (29, "...140132", "Fanny Laviqnia Lova",           13),
    (30, "...140134", "Bagoes Satria Jagad Dhita",     13),
    (31, "...140136", "Zaskia Hayatunufus",            13),
    (32, "...140142", "Wiwi Lindawati Misnadin",       11),
    (33, "...140144", "Syifa Ayudya Nurhafiza",        11),
    (34, "...140146", "Lestari Kopipah Mandasari",     9),
    (35, "...140148", "Muhammad Haykal Harsya Mevki",  8),
    (36, "...140150", "Muhammad Alif Luthfi",          13),
    (37, "...140152", "Cantika Putri Maharani",        11),
    (38, "...140160", "Amanda Marchelia C. Putri",     10),
    (39, "...140164", "Defan Atara Fahrezy",           13),
    (40, "...140166", "Hazel Ihsan Fadillah",          13),
    (41, "...140172", "Rahma Nabila Ramadhani",        10),
    (42, "...140174", "Mahira Layina Raisha Azhar",    10),
    (43, "...140176", "Abyan Hilmy Fikri Sagala",      12),
    (44, "...140178", "Diah Pertiwi Difa Nur Ilmaini", 10),
    (45, "...140180", "Muhammad Najmi Fathoni",        13),
    (46, "...140194", "Nadyatul Aulia",                10),
    (47, "...140196", "Orva Luttayya Ananditasuntoro", 12),
    (48, "...140198", "Rieska Annisa Salsabilla",      12),
    (49, "...140208", "Ananda Setiawan",               13),
    (50, "...140210", "Dira Artafirasha",              13),
]

def build_students_sheet(ws):
    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value="DAFTAR MAHASISWA — Referensi Kelompok  |  IoT Mat-D 2026")
    c.font = font(bold=True, size=12, color="FFFFFF")
    c.fill = fill("1F4E79"); c.alignment = align("center","center")
    ws.row_dimensions[1].height = 24

    hdrs = ["No","NIM","Nama Mahasiswa","No\nKelompok","Nama Kelompok"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = font(bold=True, size=10, color="FFFFFF")
        c.fill = fill("2E75B6"); c.alignment = align("center","center", wrap=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 28

    grp_map = {g[0]: g[1] for g in groups}  # {no: nama_kelompok}
    grp_fills = [
        "C6EFCE","DBEFCC","FFEB9C","FFD966","FCE4D6","F4CCCC",
        "EA9999","D5A6BD","BDD7EE","9DC3E6","C9C9C9","F2CCFF","FFE699"
    ]

    for idx, (no, nim, nama, grp_no) in enumerate(students_ref):
        r = idx + 3
        gf = fill(grp_fills[(grp_no-1) % len(grp_fills)])
        vals = [no, nim, nama, grp_no, grp_map.get(grp_no, "")]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = thin_border()
            cell.font   = font(size=10, bold=(ci==3))
            cell.fill   = gf if ci >= 4 else (fill("F7FBFF") if idx%2==0 else fill("FFFFFF"))
            cell.alignment = align("left","center") if ci==3 else align("center","center")
        ws.row_dimensions[r].height = 18

    widths = [5, 14, 30, 10, 32]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:E2"


# ─────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "Penilaian PPT-2"
build_penilaian_sheet(ws1)

ws2 = wb.create_sheet("Rubrik")
build_rubrik_sheet(ws2)

ws3 = wb.create_sheet("Daftar Mahasiswa")
build_students_sheet(ws3)

# Set default sheet
wb.active = ws1

OUT = "/projects/sandbox/sonet-product/iot-mat-d-2026/penilaian-ppt-2/PENILAIAN-PPT-2.xlsx"
wb.save(OUT)
print(f"[OK] Saved  : {OUT}")
print(f"     Sheet 1 : Penilaian PPT-2  ({len(groups)} kelompok, kolom P1–P5 + Total formula + Huruf formula)")
print(f"     Sheet 2 : Rubrik           ({len(rubrik_rows)} kriteria + skala huruf)")
print(f"     Sheet 3 : Daftar Mahasiswa ({len(students_ref)} mahasiswa + mapping kelompok)")
