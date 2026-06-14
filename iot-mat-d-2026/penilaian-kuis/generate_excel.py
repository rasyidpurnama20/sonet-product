import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# DATA KUIS 1 (49 mahasiswa)
# Kolom: No, NIM, Nama, Proyek/Fokus, K1, K2, K3, K4, K5, Total, Huruf
# ─────────────────────────────────────────────────────────────────
kuis1_rows = [
    (1,  "...120004", "Khoiriyatun",                    "Smart Gas Leak Detector",          16, 18, 12, 17, 8, 71,  "B"),
    (2,  "...120010", "Intan Tama Jessica Purba",        "Gas Leak Detector",                23, 22, 18, 16, 9, 88,  "A"),
    (3,  "...120016", "Imelda Nuris Syifa",              "Gas Leak Detector",                20, 21, 16, 11, 8, 76,  "AB"),
    (4,  "...120020", "Hentrika Aji Pamungkas",          "BJIRR Solution (banjir rob)",      22, 10,  4, 17, 8, 61,  "C"),
    (5,  "...120022", "Evlyna Fedora Hartanto",          "(proyek tidak disebut)",           23, 22, 18,  4, 8, 75,  "AB"),
    (6,  "...120028", "Nadia Bilqis",                    "(umum)",                           18, 21, 15,  5, 8, 67,  "BC"),
    (7,  "...120032", "Agnia Faradha Listya",            "Water Level Monitoring",            5, 16, 14, 16, 7, 58,  "CD"),
    (8,  "...120036", "Lusiana Kezia Arla Gracyani",     "Deteksi jalan berlubang",           6, 13, 15, 15, 7, 56,  "CD"),
    (9,  "...120040", "Muhammad Fadel Arnanda",          "EWS Banjir Rob",                   20, 21, 17, 19, 9, 86,  "A"),
    (10, "...120046", "Ramayana Sugiyo Pranoto",         "Deteksi Rob",                      20, 17, 17, 16, 8, 78,  "AB"),
    (11, "...120056", "Abdulloh Ibnu Musa",              "(cloud server)",                   16, 12,  6, 12, 6, 52,  "D"),
    (12, "...120060", "Nabila Fiesta Ramadhani",         "Absensi Pintar",                    4,  9,  5, 13, 6, 48,  "D*"),
    (13, "...120062", "Edria Filda Tsana",               "AEWG",                             21, 24, 18, 16, 9, 88,  "A"),
    (14, "...130064", "Ahmad Zidane Ainul Yaqin",        "Odor/Wound Detector",               5, 22,  5, 18, 8, 58,  "CD"),
    (15, "...130076", "Ezra Aryasatya",                  "(umum bisnis IoT)",                19, 15, 19,  8, 8, 69,  "BC"),
    (16, "...130078", "Bayhaqiy Ahmad",                  "Alarm Banjir Demak-Smg",           19, 19,  8, 18, 8, 72,  "B"),
    (17, "...130080", "Dini Permata Anisa",              "Smart Water Monitoring",            6, 15, 12, 16, 8, 57,  "CD"),
    (18, "...130082", "Achmad Fachreza Aryadewa",        "(arsitektur saja)",                 4, 18,  3,  3, 6, 34,  "E"),
    (19, "...130084", "Nimatul Karimah",                 "Accident Early Warning Grid",       4, 24,  8, 18, 9, 63,  "C"),
    (20, "...130097", "Dafa Septa Ramadhani",            "Alarm efektif",                    17, 15, 16, 13, 7, 68,  "BC"),
    (21, "...130100", "Rizki Fathiya Nur Khairunissa",   "Bersiaplah (presensi)",            18, 20, 17, 14, 9, 78,  "AB"),
    (22, "...130106", "Alya Kamila Lubna",               "Water Monitoring",                 16, 18,  5, 15, 8, 62,  "C"),
    (23, "...130110", "Ryan Anggit Nugroho",             "Railguard",                         4, 16,  6, 18, 8, 52,  "D"),
    (24, "...130112", "Sadira Najla Filzah Aryamanto",   "(sensor jarak/HC-SR04)",           24, 18, 14, 18, 9, 83,  "A"),
    (25, "...130116", "Alwan Rasyid Ramadhan",           "(arsitektur saja)",                 4, 20,  3,  3, 6, 36,  "E"),
    (26, "...130122", "Ihsan Izzat Ibrahim",             "(konseptual)",                     18, 18, 12,  4, 8, 60,  "C"),
    (27, "...130126", "Alif Alfiyansah",                 "Smart Waste System",               14, 13, 15, 16, 8, 66,  "BC"),
    (28, "...130128", "Shafana Puja Pitaloka",           "(esp/wifi)",                       17, 16,  7, 12, 7, 59,  "CD"),
    (29, "...140132", "Fanny Laviqnia Lova",             "Food Detector (IoT+AI)",            3,  5,  4,  8, 5, 25,  "E"),
    (30, "...140134", "Bagoes Satria Jagad Dhita",       "(definisi keamanan)",              16,  3,  3,  3, 5, 30,  "E"),
    (31, "...140136", "Zaskia Hayatunufus",              "Air Filter AI",                     8,  5,  5, 12, 6, 36,  "E"),
    (32, "...140142", "Wiwi Lindawati Misnadin",         "PantauKu (smartwatch)",             5, 12,  4, 17, 7, 45,  "D"),
    (33, "...140144", "Syifa Ayudya Nurhafiza",          "PantauKu (smartwatch)",            20,  6,  4, 16, 7, 53,  "D"),
    (34, "...140146", "Lestari Kopipah Mandasari",       "(kaitan proyek terpotong)",        20, 18, 17,  4, 7, 66,  "BC"),
    (35, "...140148", "Muhammad Haykal Harsya Mevki",    "Railguard",                         4, 16,  4, 17, 8, 49,  "D"),
    (36, "...140152", "Cantika Putri Maharani",          "Smartwatch child safety",          18, 10, 18, 16, 8, 70,  "B"),
    (37, "...140160", "Amanda Marchelia C. Putri",       "Smart Waste (TPS)",                 3,  4, 17, 14, 6, 44,  "E"),
    (38, "...140164", "Defan Atara Fahrezy",             "(umum)",                            4,  8,  4,  8, 6, 30,  "E"),
    (39, "...140166", "Hazel Ihsan Fadillah",            "(umum)",                            4,  8,  5,  8, 6, 31,  "E"),
    (40, "...140172", "Rahma Nabila Ramadhani",          "Deteksi sampah saluran air",        4, 17, 13, 15, 7, 56,  "CD"),
    (41, "...140174", "Mahira Layina Raisha Azhar",      "Smart TPS",                        17, 18, 16, 18, 9, 78,  "AB"),
    (42, "...140176", "Abyan Hilmy Fikri Sagala",        "(umum lengkap)",                   10, 17, 16,  6, 8, 57,  "CD"),
    (43, "...140178", "Diah Pertiwi Difa Nur Ilmaini",   "Smart Trash Detector",              8, 16, 16, 17, 8, 65,  "BC"),
    (44, "...140180", "Muhammad Najmi Fathoni",          "Smart Pollution Sensor",           15,  6, 11, 13, 6, 51,  "D"),
    (45, "...140194", "Nadyatul Aulia",                  "TPS Smart Sampah",                 14, 10, 15, 15, 7, 61,  "C"),
    (46, "...140196", "Orva Luttayya Ananditasuntoro",   "AeroSentry",                        4,  9,  8, 16, 8, 45,  "D"),
    (47, "...140198", "Rieska Annisa Salsabilla",        "AeroSentry",                        4, 12, 15, 16, 8, 55,  "CD"),
    (48, "...140208", "Ananda Setiawan",                 "(proses mini project)",             5, 10, 12,  8, 7, 42,  "E"),
    (49, "...140210", "Dira Artafirasha",                "Smart Air",                         4, 15, 10, 16, 8, 53,  "D"),
]

kuis1_comments = {
    1:  "Menyebut keamanan, arsitektur, organisasi + komponen (sensor, mikrokontroler, jaringan, aplikasi) dan mengaitkan proyek Gas Leak Detector. Kurang detail prinsip CIA & nama layer. Tingkatkan kedalaman teknis.",
    2:  "Sangat lengkap: definisi, tantangan, CIA, data privacy vs security, ancaman & pertahanan, AI/blockchain, 3-layer, TLBMC, kaitan Gas Leak Detector. Sangat baik.",
    3:  "Keamanan (perangkat low-power, jaringan publik, CIA), 3-layer, layer bisnis. Kaitan proyek masih lemah. Perjelas kaitan proyek.",
    4:  "Sangat kuat & berisi contoh nyata pada keamanan/cyber (phishing, file bajakan, WiFi publik) + kaitan proyek banjir rob. Namun tidak membahas arsitektur (layer) dan aspek bisnis. Lengkapi 2 topik tersebut.",
    5:  "Materi sangat rapi & lengkap (data security+privacy, CIA, 3-layer, TLBMC). Tidak mengaitkan dengan proyek (diminta soal). Tambahkan kaitan proyek.",
    6:  "Arsitektur 3-layer, data flow, organisasi, data security vs privacy. Tanpa kaitan proyek. Kaitkan dengan proyek.",
    7:  "Fokus pada aspek organisasi + alasan pemilihan ESP32 (proyek water level) cukup baik. Tidak ada keamanan & nama layer. Lengkapi keamanan & arsitektur.",
    8:  "Banyak membahas bisnis IoT & kaitan proyek (deteksi jalan berlubang), tetapi melewatkan keamanan & arsitektur layer; pembahasan melebar. Fokus pada materi inti pasca-UTS.",
    9:  "Sangat lengkap: arsitektur+komponen, organisasi (sosial/lingkungan/ekonomi), BMC, keamanan (CIA), kaitan proyek kuat (water level, ESP32, LoRa, Blynk, MQTT). Sangat baik.",
    10: "Keamanan (data privasi vs security, ancaman & strategi), layer & organisasi (bisnis/sosial/lingkungan), kaitan proyek Rob. Baik & menyeluruh.",
    11: "Keamanan (kategori data, kerahasiaan) cukup, tetapi arsitektur hanya \"layer sensor-proses-aplikasi\" dan tanpa aspek bisnis. Perdalam arsitektur & bisnis.",
    12: "Jawaban sangat singkat; menyebut struktur, AI, sensor & kaitan proyek absensi pintar tanpa rincian. Perlu uraian materi yang jauh lebih lengkap.",
    13: "Sangat lengkap: definisi, ancaman, defense techniques, 3-layer & 5-layer (benar), 3 aspek, kaitan proyek AEWG (panel surya). Sangat baik.",
    14: "Arsitektur 3-layer dijelaskan baik + kaitan proyek odor/wound detector sangat detail. Tanpa keamanan & bisnis. Lengkapi 2 topik.",
    15: "Kuat pada bisnis (TLBMC) & keamanan (spoofing, DoS, MITM) + 3-layer; kaitan proyek masih umum. Spesifikkan kaitan proyek.",
    16: "Arsitektur (sensor/aktuator/network/perception), data flow, data privacy & security, 3/4-layer, kaitan proyek alarm banjir. Kurang aspek bisnis. Tambah aspek bisnis.",
    17: "Pemahaman IoT & kaitan proyek (smart water monitoring) baik, tetapi arsitektur masih kabur (\"application/business layer, dll\") & keamanan minim. Perjelas layer & keamanan.",
    18: "Hanya menyebut layer arsitektur (perception/connectivity/processing/application) + definisi; sangat singkat, tanpa keamanan/bisnis/proyek. Lengkapi seluruh aspek.",
    19: "Arsitektur 5-layer dijelaskan sangat baik & dipetakan ke proyek AEWG. Tanpa keamanan & aspek bisnis (selain business layer). Lengkapi keamanan & bisnis.",
    20: "Mencakup keamanan (security vs privacy + kasus), 3-layer, 3 aspek bisnis, kaitan proyek alarm. Penyampaian singkat & banyak typo. Rapikan & perdalam.",
    21: "Keamanan (CIA), 3-layer dijelaskan, 3 aspek, definisi device, kaitan proyek presensi. Baik & lengkap.",
    22: "Definisi, sensor (HC-SR04), ESP32, CIA, 3-layer (ada keliru \"perception=mengolah data\"). Tanpa bisnis. Perbaiki definisi layer, tambah bisnis.",
    23: "Definisi + komponen + kaitan proyek Railguard sangat baik. Tanpa keamanan, layer arsitektur, & 3 aspek bisnis. Lengkapi materi inti.",
    24: "Keamanan (CIA) dijelaskan dengan penerapan pada proyek (kode sensor, autentikasi, proteksi DoS) + organisasi (HC-SR04, jaringan, router, Firebase). Sangat detail. Sangat baik.",
    25: "Hanya arsitektur 3-layer (benar) namun sangat singkat; tanpa keamanan, bisnis, proyek. Lengkapi seluruh aspek.",
    26: "Keamanan (kasus, jenis hacking ransomware/DDoS), arsitektur (3-5 layer), organisasi & alur data baik. Tanpa kaitan proyek. Tambahkan kaitan proyek.",
    27: "Komponen, keamanan (data sensor→cloud), arsitektur & organisasi bisnis (alasan beli/target), kaitan proyek smart waste. Cukup baik, perdalam CIA.",
    28: "Keamanan (tantangan + CIA) & arsitektur/organisasi (perception-network-application) + kaitan esp. Tanpa aspek bisnis & banyak typo. Tambah bisnis, rapikan.",
    29: "Membahas IoT+AI & food detector; tidak menjawab materi pasca-UTS (keamanan/arsitektur/bisnis). Jawaban melenceng dari pertanyaan.",
    30: "Hanya definisi keamanan IoT (tujuan CIA). Sangat singkat, tanpa arsitektur/bisnis/proyek. Lengkapi seluruh aspek.",
    31: "Membahas AI & tantangan secara umum + kaitan proyek air filter; mencampur AI/IoT, tanpa materi inti. Fokuskan ke materi IoT pasca-UTS.",
    32: "Definisi IoT diterapkan kuat ke proyek PantauKu, namun tidak membahas keamanan/arsitektur layer/bisnis sebagai materi. Bahas materi pasca-UTS.",
    33: "Keamanan (CIA) dijelaskan dengan penerapan proyek (GPS real-time, memori). Hanya fokus keamanan+proyek; tanpa arsitektur & bisnis. Lengkapi arsitektur & bisnis.",
    34: "Keamanan (Society 5.0, CIA), arsitektur 3-layer, organisasi (ekonomi/environment/sosial, BMC) lengkap; kaitan proyek terpotong/kosong. Lengkapi kaitan proyek.",
    35: "Definisi + arsitektur + kaitan proyek Railguard (infrared, ESP32, CCTV, gate) baik. Tanpa keamanan & bisnis. Lengkapi 2 topik.",
    36: "Keamanan (CIA, ada keliru \"keterbatasan\"), organisasi & business canvas 3 aspek + kaitan proyek smartwatch. Arsitektur layer lemah. Perbaiki istilah CIA & arsitektur.",
    37: "Hanya 3 lapisan bisnis (ekonomi/lingkungan/sosial) + kaitan proyek smart waste. Tanpa keamanan & arsitektur. Lengkapi materi inti.",
    38: "Definisi + refleksi umum tanpa rincian keamanan/layer/bisnis/proyek spesifik. Perlu uraian materi konkret.",
    39: "Umum tentang sensor/data/sistem & software/hardware/OS (sebagian keliru sebagai arsitektur IoT). Tanpa keamanan/bisnis/proyek jelas. Fokus & perdalam materi.",
    40: "Arsitektur (application/perception/network) & nilai bisnis + kaitan proyek (sensor jarak/tekanan). Tanpa keamanan. Tambahkan keamanan.",
    41: "Keamanan, arsitektur (network/sensor/hardware/cloud), organisasi bisnis 3 layer + kaitan proyek Smart TPS (sensor ultrasonik/gas/suhu). Baik & lengkap.",
    42: "Cakupan luas (pilar ekonomi/lingkungan/sosial, arsitektur, input-proses-output, organisasi, jenis hack) tetapi tanpa kaitan proyek & tanpa CIA. Kaitkan dengan proyek.",
    43: "Arsitektur (server/data/network/sensor), bisnis 3 aspek, keamanan, cost-benefit + kaitan proyek Smart Trash (ultrasonik, Firebase). Keamanan dangkal. Perdalam keamanan.",
    44: "Keamanan (data security/access/privacy) & model bisnis + kaitan proyek polusi. Tanpa arsitektur layer; banyak typo. Tambah arsitektur.",
    45: "Ancaman/kebocoran data, arsitektur & bisnis 3 layer, kaitan proyek TPS Smart Sampah. Cukup; perdalam tiap bagian.",
    46: "Refleksi AI+IoT+filtrasi & pentingnya manfaat/biaya + kaitan proyek AeroSentry. Tanpa keamanan/arsitektur layer/3 aspek. Bahas materi teknis inti.",
    47: "IoT+AI, arsitektur sesuai kebutuhan, organisasi bisnis & analisis biaya + kaitan proyek AeroSentry. Tanpa keamanan & nama layer. Tambah keamanan & layer.",
    48: "Menyebut tahapan mini project (ide/input/proses/output, pihak, biaya, arsitektur, organisasi) tetapi tanpa rincian & nama proyek. Perlu substansi materi.",
    49: "Fokus proyek Smart Air + komponen arsitektur (MQ135, ESP32, NodeMCU, ESP8266, TDS) baik. Tanpa keamanan & 3 aspek bisnis/nama layer. Lengkapi materi inti.",
}

# ─────────────────────────────────────────────────────────────────
# DATA KUIS 2 (50 mahasiswa)
# Kolom: No, NIM, Nama, Ide/Judul, D1, D2, D3, D4, D5, Total, Huruf
# ─────────────────────────────────────────────────────────────────
kuis2_rows = [
    (1,  "...120004", "Khoiriyatun",                    "Monitoring listrik via HP",              15, 16,  8, 16,  8, 63,  "C"),
    (2,  "...120010", "Intan Tama Jessica Purba",        "(mengacu Gas Leak Detector)",           14, 24, 14, 17,  9, 78,  "AB"),
    (3,  "...120016", "Imelda Nuris Syifa",              "Monitoring + auto-off",                 17, 25, 14, 17,  9, 82,  "A"),
    (4,  "...120020", "Hentrika Aji Pamungkas",          "Penghangat Ruangan Pintar",             17, 23, 10, 15,  9, 74,  "B"),
    (5,  "...120022", "Evlyna Fedora Hartanto",          "Electricity Smart Home (panel surya)",  17, 16, 11, 15,  8, 67,  "BC"),
    (6,  "...120028", "Nadia Bilqis",                    "Deteksi + auto-off perangkat",          16, 22, 14, 15,  9, 76,  "AB"),
    (7,  "...120032", "Agnia Faradha Listya",            "Monitoring + notifikasi",               17, 24, 14, 18,  9, 82,  "A"),
    (8,  "...120036", "Lusiana Kezia Arla Gracyani",     "Monitoring via HP",                     15, 15,  7, 15,  7, 59,  "CD"),
    (9,  "...120040", "Muhammad Fadel Arnanda",          "Gateway per perangkat",                 16, 23, 14, 18,  9, 80,  "A"),
    (10, "...120046", "Ramayana Sugiyo Pranoto",         "Arus Hemat",                            17, 22, 12, 17,  9, 77,  "AB"),
    (11, "...120056", "Abdulloh Ibnu Musa",              "Monitoring sekring/perangkat",          16, 20, 13, 15,  8, 72,  "B"),
    (12, "...120060", "Nabila Fiesta Ramadhani",         "Kontrol TV/AC via app",                 14, 13, 10, 13,  7, 57,  "CD"),
    (13, "...120062", "Edria Filda Tsana",               "Monitoring + Blynk",                    18, 26, 12, 17,  9, 82,  "A"),
    (14, "...130064", "Ahmad Zidane Ainul Yaqin",        "Kontrol perangkat rumah",               18, 20,  4, 13,  8, 63,  "C"),
    (15, "...130076", "Ezra Aryasatya",                  "Monitoring + kontrol suara",            18, 26, 14, 16,  9, 83,  "A"),
    (16, "...130078", "Bayhaqiy Ahmad",                  "Sensor per perangkat",                  17, 24, 15, 17,  9, 82,  "A"),
    (17, "...130080", "Dini Permata Anisa",              "Monitoring antar-perangkat",            17, 22, 10, 14,  8, 71,  "B"),
    (18, "...130082", "Achmad Fachreza Aryadewa",        "Deteksi arus masuk/keluar",             17, 27, 13, 16,  9, 82,  "A"),
    (19, "...130084", "Nimatul Karimah",                 "Monitor rumah (solar)",                 15, 24, 16, 15,  9, 79,  "AB"),
    (20, "...130097", "Dafa Septa Ramadhani",            "Kipas auto-off",                        14, 15,  7, 11,  7, 54,  "D"),
    (21, "...130100", "Rizki Fathiya Nur Khairunissa",   "Auto AC/lampu + monitor",               18, 25, 12, 12,  9, 76,  "AB"),
    (22, "...130106", "Alya Kamila Lubna",               "Deteksi on/off + kontrol",              16, 25, 14, 14,  9, 78,  "AB"),
    (23, "...130110", "Ryan Anggit Nugroho",             "Monitor kWh dari meteran",              17, 20,  4, 17,  8, 66,  "BC"),
    (24, "...130112", "Sadira Najla Filzah Aryamanto",   "Monitoring + BMC lengkap",              17, 24, 15, 19,  9, 84,  "A"),
    (25, "...130116", "Alwan Rasyid Ramadhan",           "Smart TV auto-off",                     13, 16,  3, 10,  7, 49,  "D"),
    (26, "...130122", "Ihsan Izzat Ibrahim",             "Monitor pergerakan arus",               15, 18, 12, 15,  8, 68,  "BC"),
    (27, "...130126", "Alif Alfiyansah",                 "Hemat listrik + bill realtime",         16, 24, 12, 16,  9, 77,  "AB"),
    (28, "...130128", "Shafana Puja Pitaloka",           "Monitor dekat token listrik",           14, 20,  4, 15,  8, 61,  "C"),
    (29, "...140132", "Fanny Laviqnia Lova",             "Pendingin Udara Otomatis",              14, 14,  3, 10,  7, 48,  "D"),
    (30, "...140134", "Bagoes Satria Jagad Dhita",       "Sensor per alat + MQTT",                17, 24, 15, 16,  9, 81,  "A"),
    (31, "...140136", "Zaskia Hayatunufus",              "Watt Watch (kos)",                      15, 13,  5, 12,  7, 52,  "D"),
    (32, "...140142", "Wiwi Lindawati Misnadin",         "Monitor + cutoff + biometrik",          18, 27, 17, 18, 10, 90,  "A"),
    (33, "...140144", "Syifa Ayudya Nurhafiza",          "Auto-off + alarm watt",                 16, 23, 12, 15,  9, 75,  "AB"),
    (34, "...140146", "Lestari Kopipah Mandasari",       "Smart Meter Surya Panel",               17, 20, 13, 19,  9, 78,  "AB"),
    (35, "...140148", "Muhammad Haykal Harsya Mevki",    "Sensor arus + infrared",                12, 18, 12, 14,  7, 63,  "C"),
    (36, "...140150", "Muhammad Alif Luthfi",            "Monitoring + cloud + app",              16, 22, 17, 15,  9, 79,  "AB"),
    (37, "...140152", "Cantika Putri Maharani",          "Auto-off + AI/IoT",                     16, 12, 12, 18,  8, 66,  "BC"),
    (38, "...140160", "Amanda Marchelia C. Putri",       "Hemat listrik (perumahan elit)",        16, 10,  8, 16,  7, 57,  "CD"),
    (39, "...140164", "Defan Atara Fahrezy",             "Smart Silent Room (kebisingan)",        10, 16,  4, 11,  7, 48,  "D"),
    (40, "...140166", "Hazel Ihsan Fadillah",            "Smart Door Lock",                        8, 15, 12, 12,  7, 54,  "D"),
    (41, "...140172", "Rahma Nabila Ramadhani",          "Monitor berbasis suhu",                 12, 16,  4, 13,  7, 52,  "D"),
    (42, "...140174", "Mahira Layina Raisha Azhar",      "Aplikasi monitoring",                   18, 10, 17, 15,  9, 69,  "BC"),
    (43, "...140176", "Abyan Hilmy Fikri Sagala",        "Blynk + auto-off (kampus)",             17, 26, 14, 16,  9, 82,  "A"),
    (44, "...140178", "Diah Pertiwi Difa Nur Ilmaini",   "Listrik pintar + notifikasi",           16, 20, 12, 16,  8, 72,  "B"),
    (45, "...140180", "Muhammad Najmi Fathoni",          "Konsep-arsitektur-bisnis",              15, 18, 12, 16,  8, 69,  "BC"),
    (46, "...140194", "Nadyatul Aulia",                  "Anti-boros listrik",                    16, 23, 12, 16,  9, 76,  "AB"),
    (47, "...140196", "Orva Luttayya Ananditasuntoro",   "Chairzz (anti-ngantuk)",                 8, 14,  5, 12,  7, 46,  "D"),
    (48, "...140198", "Rieska Annisa Salsabilla",        "Monitor + cloud AI",                    16, 20, 13, 16,  8, 73,  "B"),
    (49, "...140208", "Ananda Setiawan",                 "Monitor + notifikasi + remote",         18, 28, 17, 18, 10, 91,  "A"),
    (50, "...140210", "Dira Artafirasha",                "Monitoring real-time",                  15, 22, 11, 16,  8, 72,  "B"),
]

kuis2_comments = {
    1:  "Ide & bisnis (3 aspek) jelas, komponen disebut (sensor arus, mikrokontroler, WiFi, app). Arsitektur dangkal & keamanan lemah (hanya notifikasi). Perdalam arsitektur & mekanisme keamanan nyata.",
    2:  "Teknis sangat lengkap (layer, sensor, alur, bisnis 3 aspek, keamanan, AI). Namun uraian mengacu ke Gas Leak Detector, bukan murni energy monitoring. Sesuaikan dengan tema soal.",
    3:  "Sangat lengkap & terstruktur: ide, arsitektur berlapis (ESP32, MQTT, cloud), keamanan (enkripsi+autentikasi), bisnis (HaaS, integrasi pembayaran). Sangat baik.",
    4:  "\"Penghangat Ruangan Pintar\" relevan (hemat energi); detail sensor suhu+arus, ESP WiFi, Blynk, batas daya, alur data, bisnis. Keamanan masih umum. Perdalam keamanan.",
    5:  "Ide bagus (data tiap 3 ms, deteksi konsleting, panel surya) & bisnis tertarget. Arsitektur kurang komponen/layer; keamanan agak umum. Rinci komponen & layer.",
    6:  "Lengkap: ide auto-off, arsitektur (sensor→ESP32→WiFi→app), keamanan (password+enkripsi DB), bisnis. Baik.",
    7:  "Sangat lengkap: ide, arsitektur (ESP32, sensor arus & tegangan, kontrol on/off), segmentasi pasar tajam. Catatan: MITM/firmware tampering itu ancaman, bukan metode keamanan. Perbaiki konsep keamanan.",
    8:  "Ide & bisnis (usul pemerintah, lapangan kerja) panjang, tetapi arsitektur dangkal & keamanan diakui \"sulit dijamin\". Bertele-tele. Perkuat arsitektur & keamanan.",
    9:  "Lengkap & kuat: gateway per perangkat, alur jelas, keamanan (MQTT/HTTP→CIA), bisnis BMC (key resource/partner/revenue/customer). Sangat baik.",
    10: "\"Arus Hemat\": ide & arsitektur (ESP32, cloud besar, output harian/bulanan/tahunan + biaya) baik, bisnis tertarget. Keamanan cukup. Baik.",
    11: "Ide detail (sensor di sekring/kabel), alur lengkap, keamanan (application+network layer, sandi), bisnis berkembang (rumah→industri). Baik, sebut sensor/MCU spesifik.",
    12: "Mencakup keempat aspek (TV/AC, sensor arus, WiFi, sidik jari) namun ringkas; arsitektur tanpa kedalaman. Perinci arsitektur & alur data.",
    13: "Sangat lengkap: alur sensor→gateway(filter)→cloud→Blynk, layer dinamai, keamanan (auth Blynk), bisnis 3 aspek. Sangat baik.",
    14: "Ide naratif kuat & arsitektur baik (sensor per alat, data 0/1, MCU→app). Tidak ada bagian keamanan. Tambahkan keamanan.",
    15: "Canggih: kontrol suara + monitoring, layer (perception/network/processing+AI), keamanan (scan aktivitas mencurigakan), bisnis (value/langganan/resource). Sangat baik.",
    16: "Lengkap: arsitektur 3-layer (4G/5G/WiFi), keamanan (diskusi WiFi global vs lokal, malware), bisnis tertarget urban. Sangat baik.",
    17: "Ide fleksibel multi-perangkat, layer dinamai (networking/perception/application), bisnis (ekonomi/lingkungan). Keamanan (cloud Google) dangkal. Perdalam keamanan.",
    18: "Sangat terstruktur 4 bagian; layer dinamai tepat (perception/connectivity/processing/application), keamanan (firewall), bisnis (mitra developer rumah). Sangat baik.",
    19: "Arsitektur 3-layer (LoRaWAN) & keamanan CIA dijelaskan baik, bisnis (paket+premium). Ide sedikit bergeser ke pemantauan/CCTV. Fokuskan ke energy monitoring.",
    20: "Ide kipas auto-off relevan, tetapi salah memetakan layer (\"application layer = sensor suhu\") & keamanan minim. Perbaiki pemetaan layer & keamanan.",
    21: "Sangat detail: skenario AC/lampu otomatis, sensor panas & arus, ESP32, relay, alur, biaya ~500rb. Aspek bisnis tipis (hanya biaya). Lengkapi model bisnis.",
    22: "Lengkap & berlapis: layer dinamai, keamanan (sidik jari, data ke app pemilik), bisnis (pihak terlibat). Baik sekali.",
    23: "Ide (kWh dari meteran) & arsitektur (ESP32, cloud, batas harian + notif) bagus, bisnis & keterbatasan dibahas. Tanpa bagian keamanan. Tambahkan keamanan.",
    24: "Arsitektur, keamanan, dan Business Model Canvas lengkap (channels, revenue, cost, target, jasa install/service). Sangat baik.",
    25: "Ide smart TV auto-off (relevan hemat energi) & arsitektur (kamera ESP32, kontroler, app). Tanpa keamanan; bisnis tipis. Lengkapi keamanan & bisnis.",
    26: "Ide & arsitektur (Arduino, sensor, cloud, laporan otomatis), keamanan (update versi, cybersecurity), bisnis (murah, UI simpel, mitra). Cukup baik.",
    27: "Lengkap: ide hemat+bill realtime, arsitektur (sensor arus/tegangan, MCU, WiFi, cloud, app), keamanan (autentikasi), bisnis (per sistem/bundle rumah/perusahaan). Baik sekali.",
    28: "Arsitektur, alur input-proses-output & bisnis (produsen, online store, ~600rb, lingkungan) baik. Tanpa bagian keamanan. Tambahkan keamanan.",
    29: "Ide AC otomatis relevan, arsitektur cukup (MCU sebagai otak) namun tanpa komponen/layer spesifik & tanpa keamanan. Lengkapi komponen & keamanan.",
    30: "Lengkap & matang: sensor per alat, MQTT, cloud, keamanan (login+enkripsi+WiFi aman), bisnis (paket+app, target rumah/kos). Sangat baik.",
    31: "\"Watt Watch\" untuk kos (ide kontekstual bagus), tetapi arsitektur dangkal & keamanan disalahartikan (pembagian tagihan). Perkuat arsitektur & keamanan.",
    32: "Sangat lengkap: arsitektur 4-lapis (ESP32, gateway, broker MQTT, app), keamanan kuat (enkripsi MQTT, token/OTP, biometrik), bisnis rinci (margin 40%, garansi). Istimewa.",
    33: "Lengkap & berlapis: sensor kehadiran + alarm watt, layer dinamai, keamanan (kerahasiaan/ketersediaan), bisnis (mitra energi terbarukan). Baik sekali.",
    34: "\"Smart Meter Surya Panel\": layer, keamanan CIA, dan BMC lengkap (key activities/partners PLN/value/channels/segment). Catatan: \"integrasi\" pada CIA seharusnya \"integritas\". Sangat baik.",
    35: "Arsitektur (sensor arus + infrared kehadiran), keamanan (password+proteksi data), bisnis (target & manfaat). Tanpa intro ide & nama layer. Tambah ide & layer.",
    36: "Terstruktur 4 bagian; keamanan sangat kuat (HTTPS/TLS, password, device ID unik, pembatasan akses), bisnis (penjualan+premium). Baik sekali.",
    37: "Ide & bisnis (3 aspek detail) kuat, tetapi arsitektur keliru (application layer = rumah; tanpa perception/sensor). Perbaiki pemetaan layer.",
    38: "Ide & bisnis (target perumahan elit) baik, tetapi arsitektur lemah (tanpa komponen/layer) & keamanan dangkal (\"WiFi sendiri = aman\"). Perkuat arsitektur & keamanan.",
    39: "\"Smart Silent Room\" (deteksi kebisingan) di luar tema energy monitoring; arsitektur (sensor suara→ESP32→server) cukup, keamanan minim. Sesuaikan tema soal.",
    40: "\"Smart Door Lock\" menyimpang dari tema; terstruktur (Arduino, microservo, password) namun bukan energy monitoring. Sesuaikan tema soal.",
    41: "Fokus bergeser ke suhu/panel surya (kurang ke konsumsi listrik); arsitektur cukup, tanpa keamanan. Fokuskan ke monitoring energi & tambah keamanan.",
    42: "Ide aplikasi rinci & keamanan sangat baik (login + password+pin, alert login device lain). Namun arsitektur lemah (tanpa sensor/MCU/layer). Lengkapi arsitektur perangkat keras.",
    43: "Sangat konkret: Blynk, sensor listrik/LDR/ESP32, if-else, auto-off, API keys, alur lengkap, peluang pasar kampus. Sangat baik.",
    44: "Lengkap: ide (sensor arus, iCloud, notifikasi), arsitektur (sensor/network/data/service), keamanan, bisnis 3 aspek. Baik.",
    45: "Metodis: konsep→latar→desain I/P/O→arsitektur 3-layer→keamanan (security/privacy)→bisnis (target, 24/7, cost, end user). Cukup baik, perinci komponen.",
    46: "Lengkap: latar masalah, arsitektur 3-layer (sensor arus, MCU, WiFi, app), keamanan (login), bisnis (jual/jasa install). Baik sekali.",
    47: "\"Chairzz\" (kursi anti-ngantuk) di luar tema; struktur ada tetapi keamanan disalahartikan (efek getar). Sesuaikan tema soal.",
    48: "Lengkap: ide kontrol listrik, arsitektur (sensor→MCU→internet→app), keamanan (enkripsi + sensor overload), bisnis (premium, perusahaan). Baik.",
    49: "Paling lengkap: ide jelas, arsitektur rinci (SCT-013, ESP32, relay, router; firmware/cloud/dashboard), keamanan (email-password, enkripsi, akses, backup), bisnis (target + model + mitra PLN). Istimewa.",
    50: "Lengkap & ringkas: arsitektur berkomponen (sensor arus, ESP32/NodeMCU/ESP8266, WiFi, app), keamanan (cyber security), bisnis (masyarakat/pemerintah). Baik.",
}

# ─────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────

def grade_color(huruf):
    """Return hex fill color based on letter grade."""
    mapping = {
        "A":  "C6EFCE",   # green
        "AB": "DBEFCC",   # light green
        "B":  "FFEB9C",   # yellow
        "BC": "FFD966",   # gold
        "C":  "FCE4D6",   # light orange
        "CD": "F4CCCC",   # light red
        "D":  "EA9999",   # red
        "D*": "EA9999",   # red
        "E":  "CC0000",   # dark red
    }
    return mapping.get(huruf, "FFFFFF")

def grade_font_color(huruf):
    return "FFFFFF" if huruf == "E" else "000000"

def thin_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:    cell.font = font
    if fill:    cell.fill = fill
    if align:   cell.alignment = align
    if border:  cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def build_sheet(ws, title_text, headers, data_rows, comments_dict,
                score_cols,          # list of 0-based indices of score columns in headers
                total_col_idx,       # 0-based index of Total column
                huruf_col_idx):      # 0-based index of Huruf column

    # ── colour palette ──────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
    TITLE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    SUBHDR_FILL = PatternFill("solid", fgColor="D6E4F0")
    SUBHDR_FONT = Font(bold=True, color="1F4E79", name="Calibri", size=10)
    ALT_FILL   = PatternFill("solid", fgColor="F2F7FC")
    DEFAULT_FONT = Font(name="Calibri", size=10)
    TOTAL_FONT   = Font(bold=True, name="Calibri", size=10)
    COMMENT_FILL = PatternFill("solid", fgColor="FDFEFE")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right_align  = Alignment(horizontal="right",  vertical="center")

    bdr = thin_border()

    n_cols = len(headers) + 1   # +1 for Komentar

    # ── Row 1: merged title ──────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t_cell = ws.cell(row=1, column=1, value=title_text)
    t_cell.font    = TITLE_FONT
    t_cell.fill    = TITLE_FILL
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: bobot sub-header for score columns ────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    subtitle = ws.cell(row=2, column=1, value="IoT Mat-D 2026  |  Skala: A≥80 · AB 75–79 · B 70–74 · BC 65–69 · C 60–64 · CD 55–59 · D 45–54 · E<45")
    subtitle.font  = Font(italic=True, color="444444", name="Calibri", size=10)
    subtitle.fill  = PatternFill("solid", fgColor="EBF3FB")
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: column headers ────────────────────────────────────
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = center_align
        c.border    = bdr
    # Komentar header
    c = ws.cell(row=3, column=n_cols, value="Komentar / Catatan")
    c.font      = HDR_FONT
    c.fill      = HDR_FILL
    c.alignment = center_align
    c.border    = bdr
    ws.row_dimensions[3].height = 30

    # ── Data rows ────────────────────────────────────────────────
    for r_idx, row in enumerate(data_rows):
        excel_row = r_idx + 4
        no = row[0]
        huruf = str(row[huruf_col_idx])  # 0-based: same index as headers

        row_fill = ALT_FILL if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=excel_row, column=ci)
            c.value     = val
            c.border    = bdr
            c.alignment = center_align if ci != 3 else Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Score columns → use number format
            if (ci - 1) in score_cols:
                c.font = DEFAULT_FONT
                c.fill = row_fill
                c.number_format = "0"
            elif (ci - 1) == total_col_idx:
                # Total column
                c.font = TOTAL_FONT
                g_fill = PatternFill("solid", fgColor=grade_color(huruf))
                c.fill = g_fill
                c.font = Font(bold=True, color=grade_font_color(huruf), name="Calibri", size=10)
            elif (ci - 1) == huruf_col_idx:
                # Huruf column
                g_fill = PatternFill("solid", fgColor=grade_color(huruf))
                c.fill = g_fill
                c.font = Font(bold=True, color=grade_font_color(huruf), name="Calibri", size=10)
            else:
                c.font = DEFAULT_FONT
                c.fill = row_fill

        # Komentar column
        comment_text = comments_dict.get(no, "")
        kc = ws.cell(row=excel_row, column=n_cols, value=comment_text)
        kc.font      = Font(name="Calibri", size=10, italic=True)
        kc.fill      = COMMENT_FILL
        kc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        kc.border    = bdr

        ws.row_dimensions[excel_row].height = 55

    # ── Column widths ────────────────────────────────────────────
    col_widths = []
    for ci, h in enumerate(headers, start=1):
        if h == "No":           col_widths.append((ci, 5))
        elif h == "NIM":        col_widths.append((ci, 13))
        elif h == "Nama":       col_widths.append((ci, 28))
        elif h in ("Proyek/Fokus", "Ide/Judul"):
                                col_widths.append((ci, 28))
        elif h in ("Total",):   col_widths.append((ci, 8))
        elif h in ("Huruf",):   col_widths.append((ci, 7))
        elif h.startswith("K") or h.startswith("D"):
                                col_widths.append((ci, 6))
        else:                   col_widths.append((ci, 12))
    col_widths.append((n_cols, 70))   # Komentar

    for ci, w in col_widths:
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Freeze header rows ───────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on header row ────────────────────────────────
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"

    return ws


# ─────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Sheet K (Kuis 1)
ws_k = wb.active
ws_k.title = "Kuis 1 (K)"
headers_k = ["No", "NIM", "Nama", "Proyek/Fokus", "K1", "K2", "K3", "K4", "K5", "Total", "Huruf"]
# score_cols (0-based in headers_k): K1=4, K2=5, K3=6, K4=7, K5=8
# total_col_idx = 9, huruf_col_idx = 10
build_sheet(ws_k,
    title_text    = "PENILAIAN KUIS 1 — Materi Pasca-UTS & Kaitan Proyek  |  IoT Mat-D 2026",
    headers       = headers_k,
    data_rows     = kuis1_rows,
    comments_dict = kuis1_comments,
    score_cols    = [4, 5, 6, 7, 8],
    total_col_idx = 9,
    huruf_col_idx = 10)

# Sheet D (Kuis 2)
ws_d = wb.create_sheet(title="Kuis 2 (D)")
headers_d = ["No", "NIM", "Nama", "Ide/Judul", "D1", "D2", "D3", "D4", "D5", "Total", "Huruf"]
# score_cols (0-based): D1=4, D2=5, D3=6, D4=7, D5=8
# total_col_idx = 9, huruf_col_idx = 10
build_sheet(ws_d,
    title_text    = "PENILAIAN KUIS 2 — Desain Smart Home Energy Monitoring  |  IoT Mat-D 2026",
    headers       = headers_d,
    data_rows     = kuis2_rows,
    comments_dict = kuis2_comments,
    score_cols    = [4, 5, 6, 7, 8],
    total_col_idx = 9,
    huruf_col_idx = 10)

# ── Save ─────────────────────────────────────────────────────────
out_path = "/projects/sandbox/sonet-product/iot-mat-d-2026/penilaian-kuis/PENILAIAN-KUIS-MAHASISWA.xlsx"
wb.save(out_path)
print(f"[OK] Saved: {out_path}")
print(f"     Sheet 'Kuis 1 (K)' : {len(kuis1_rows)} baris + komentar")
print(f"     Sheet 'Kuis 2 (D)' : {len(kuis2_rows)} baris + komentar")
